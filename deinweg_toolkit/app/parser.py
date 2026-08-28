"""Einlesen von Zeitlisten aus xlsx/xls/csv in ein einheitliches Format.

Zielformat pro Zeile:
    {datum: 'YYYY-MM-DD', start: 'HH:MM'|None, ende: 'HH:MM'|None,
     klient: str, beschreibung: str, dauer_min: int,
     mitarbeiter: str, warnung: str|None}
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import re
import unicodedata

from openpyxl import load_workbook

# --- Spaltenerkennung -------------------------------------------------------

# Reihenfolge = Priorität. Exakte Treffer schlagen Teiltreffer.
ALIASES: dict[str, list[str]] = {
    "datum": ["tag", "datum", "date", "day", "kalendertag"],
    "start": ["start", "beginn", "von", "anfang", "startzeit",
              "start time", "uhrzeit von", "beginnt"],
    "ende": ["ende", "end", "bis", "endzeit", "end time",
             "uhrzeit bis", "endet"],
    "dauer": ["dauer", "duration", "zeitaufwand", "arbeitszeit",
              "stunden", "hours", "menge"],
    "klient": ["aufgabe", "task", "klient", "klientin", "kunde", "client",
               "betreute person", "betreuter", "bewohner", "projekt",
               "leistungsempfaenger", "name"],
    "mitarbeiter": ["tags", "mitarbeiter", "mitarbeiterin", "employee",
                    "betreuer", "betreuerin", "fachkraft", "user",
                    "bearbeiter", "kuerzel"],
    "beschreibung": ["beschreibung arbeitseinheit", "beschreibung", "notiz",
                     "note", "notes", "taetigkeit", "kommentar",
                     "description", "bemerkung", "leistung", "inhalt"],
}

FUSSZEILEN = {"gesamt", "total", "summe", "sum", "gesamtsumme", "insgesamt"}

# Werte in der Klienten-Spalte, die keine betreute Person sind
NICHT_ABRECHENBAR = {"sonstiges", "sonstige", "intern", "team", "-", ""}


def norm(text) -> str:
    """Kleinschreibung, Umlaute aufgeloest, Sonderzeichen weg."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = (s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
           .replace("ß", "ss"))
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def mappe_spalten(header: list) -> dict[str, int]:
    """Ordnet Spaltenindizes den logischen Feldern zu."""
    kopf = [norm(h) for h in header]
    mapping: dict[str, int] = {}

    # 1. Durchgang: exakte Alias-Treffer
    for feld, aliase in ALIASES.items():
        for idx, h in enumerate(kopf):
            if h and h in aliase and idx not in mapping.values():
                mapping[feld] = idx
                break

    # 2. Durchgang: Teiltreffer fuer noch fehlende Felder
    for feld, aliase in ALIASES.items():
        if feld in mapping:
            continue
        for idx, h in enumerate(kopf):
            if not h or idx in mapping.values():
                continue
            if any(a in h or h in a for a in aliase):
                mapping[feld] = idx
                break
    return mapping


def finde_kopfzeile(zeilen: list[list]) -> int:
    """Sucht in den ersten 15 Zeilen die Zeile mit den meisten Spaltentreffern."""
    bester, beste_punkte = 0, 0
    for i, zeile in enumerate(zeilen[:15]):
        treffer = len(mappe_spalten(zeile))
        if treffer > beste_punkte:
            bester, beste_punkte = i, treffer
    if beste_punkte < 2:
        raise ValueError(
            "Keine Kopfzeile erkannt. Erwartet werden Spalten wie "
            "Datum/Tag, Start, Ende, Dauer, Aufgabe/Klient."
        )
    return bester


# --- Wertkonvertierung ------------------------------------------------------

DATUMSFORMATE = ["%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y",
                 "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]


def parse_datum(wert):
    if wert is None or wert == "":
        return None
    if isinstance(wert, dt.datetime):
        return wert.date()
    if isinstance(wert, dt.date):
        return wert
    s = str(wert).strip()
    # fuehrenden Wochentag entfernen ("Mo, 06.04.2026")
    s = re.sub(r"^[A-Za-zÄÖÜäöü]{2,10}\.?,?\s+", "", s)
    s = s.split(" ")[0] if " " in s and ":" in s else s
    for fmt in DATUMSFORMATE:
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_zeit(wert):
    if wert is None or wert == "":
        return None
    if isinstance(wert, dt.datetime):
        return wert.strftime("%H:%M")
    if isinstance(wert, dt.time):
        return wert.strftime("%H:%M")
    s = str(wert).strip().upper()
    m = re.search(r"(\d{1,2})[:.](\d{2})", s)
    if not m:
        return None
    stunde, minute = int(m.group(1)), int(m.group(2))
    if "PM" in s and stunde < 12:
        stunde += 12
    if "AM" in s and stunde == 12:
        stunde = 0
    if not (0 <= stunde <= 23 and 0 <= minute <= 59):
        return None
    return f"{stunde:02d}:{minute:02d}"


def parse_dauer(wert) -> int | None:
    """Gibt Minuten zurueck. Versteht '02:00', '1,5', '90 min', Excel-Zeitwerte."""
    if wert is None or wert == "":
        return None
    if isinstance(wert, dt.timedelta):
        return int(round(wert.total_seconds() / 60))
    if isinstance(wert, dt.datetime):
        return wert.hour * 60 + wert.minute
    if isinstance(wert, dt.time):
        return wert.hour * 60 + wert.minute
    if isinstance(wert, (int, float)):
        # Excel-Zeit ist ein Tagesbruchteil, sonst Dezimalstunden
        return int(round(wert * 1440)) if 0 < wert < 1 else int(round(wert * 60))
    s = str(wert).strip().lower().replace(" ", "")
    m = re.match(r"^(-?\d{1,3})[:.](\d{2})(?::\d{2})?$", s)
    if m:
        vorz = -1 if m.group(1).startswith("-") else 1
        return vorz * (abs(int(m.group(1))) * 60 + int(m.group(2)))
    if s.endswith("min"):
        s = s[:-3]
        try:
            return int(round(float(s.replace(",", "."))))
        except ValueError:
            return None
    s = s.replace("h", "").replace("std", "")
    try:
        return int(round(float(s.replace(",", ".")) * 60))
    except ValueError:
        return None


def dauer_aus_spanne(start: str | None, ende: str | None) -> int | None:
    if not start or not ende:
        return None
    sh, sm = (int(x) for x in start.split(":"))
    eh, em = (int(x) for x in ende.split(":"))
    minuten = (eh * 60 + em) - (sh * 60 + sm)
    if minuten < 0:            # ueber Mitternacht
        minuten += 24 * 60
    return minuten


def hhmm(minuten: int) -> str:
    vorz = "-" if minuten < 0 else ""
    minuten = abs(int(minuten))
    return f"{vorz}{minuten // 60:02d}:{minuten % 60:02d}"


def fingerprint(zeile: dict) -> str:
    roh = "|".join([
        norm(zeile["mitarbeiter"]), zeile["datum"], zeile["start"] or "",
        zeile["ende"] or "", norm(zeile["klient"]),
        norm(zeile["beschreibung"]), str(zeile["dauer_min"]),
    ])
    return hashlib.sha1(roh.encode("utf-8")).hexdigest()


# --- Rohdaten lesen ---------------------------------------------------------

def lies_xlsx(daten: bytes) -> list[list]:
    wb = load_workbook(io.BytesIO(daten), data_only=True, read_only=True)
    ws = wb.active
    zeilen = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return zeilen


def lies_csv(daten: bytes) -> list[list]:
    text = None
    for kodierung in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = daten.decode(kodierung)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Zeichenkodierung der CSV-Datei nicht lesbar.")
    probe = "\n".join(text.splitlines()[:10])
    try:
        dialekt = csv.Sniffer().sniff(probe, delimiters=",;\t|")
        trenner = dialekt.delimiter
    except csv.Error:
        trenner = ";" if probe.count(";") > probe.count(",") else ","
    return [r for r in csv.reader(io.StringIO(text), delimiter=trenner)]


# --- Hauptfunktion ----------------------------------------------------------

def lies_datei(dateiname: str, daten: bytes, mitarbeiter_fallback: str = "",
               mitarbeiter_erzwingen: bool = False) -> tuple[list[dict], dict]:
    endung = dateiname.lower().rsplit(".", 1)[-1]
    if endung in ("xlsx", "xlsm"):
        rohzeilen = lies_xlsx(daten)
    elif endung == "csv":
        rohzeilen = lies_csv(daten)
    else:
        raise ValueError(f"Dateityp .{endung} wird nicht unterstuetzt "
                         "(erlaubt: .xlsx, .xlsm, .csv).")

    if not rohzeilen:
        raise ValueError("Die Datei enthaelt keine Zeilen.")

    kopf_idx = finde_kopfzeile(rohzeilen)
    mapping = mappe_spalten(rohzeilen[kopf_idx])
    if "datum" not in mapping:
        raise ValueError("Keine Datumsspalte gefunden (erwartet 'Tag' oder 'Datum').")

    def hole(zeile, feld):
        idx = mapping.get(feld)
        if idx is None or idx >= len(zeile):
            return None
        return zeile[idx]

    ergebnis: list[dict] = []
    statistik = {"gelesen": 0, "uebersprungen": 0, "kopfzeile": kopf_idx + 1,
                 "spalten": {k: rohzeilen[kopf_idx][v] for k, v in mapping.items()}}

    for zeile in rohzeilen[kopf_idx + 1:]:
        if not any(z not in (None, "") for z in zeile):
            continue
        erste = norm(zeile[0]) if zeile else ""
        if erste in FUSSZEILEN:
            continue

        datum = parse_datum(hole(zeile, "datum"))
        if datum is None:
            statistik["uebersprungen"] += 1
            continue

        start = parse_zeit(hole(zeile, "start"))
        ende = parse_zeit(hole(zeile, "ende"))
        dauer = parse_dauer(hole(zeile, "dauer"))
        warnungen = []

        if dauer is None:
            dauer = dauer_aus_spanne(start, ende)
            if dauer is not None:
                warnungen.append("Dauer aus Start/Ende berechnet")
        if dauer is None:
            statistik["uebersprungen"] += 1
            continue
        if dauer <= 0:
            warnungen.append("Dauer ist 0 oder negativ")
        if dauer > 12 * 60:
            warnungen.append("Dauer ueber 12 Stunden")

        spanne = dauer_aus_spanne(start, ende)
        if spanne is not None and abs(spanne - dauer) > 2:
            warnungen.append(
                f"Dauer ({hhmm(dauer)}) passt nicht zu Start/Ende ({hhmm(spanne)})")

        klient = str(hole(zeile, "klient") or "").strip() or "Ohne Zuordnung"
        beschreibung = str(hole(zeile, "beschreibung") or "").strip()

        mitarbeiter = str(hole(zeile, "mitarbeiter") or "").strip()
        if mitarbeiter_erzwingen or not mitarbeiter:
            mitarbeiter = mitarbeiter_fallback.strip() or mitarbeiter
        if not mitarbeiter:
            raise ValueError(
                "Kein Mitarbeiter erkennbar. Bitte im Formular einen Namen "
                "eintragen oder eine Spalte 'Tags'/'Mitarbeiter' ergaenzen."
            )

        eintrag = {
            "datum": datum.isoformat(),
            "monat": datum.strftime("%Y-%m"),
            "start": start,
            "ende": ende,
            "klient": klient,
            "beschreibung": beschreibung,
            "dauer_min": int(dauer),
            "mitarbeiter": mitarbeiter,
            "abrechenbar": 0 if norm(klient) in NICHT_ABRECHENBAR else 1,
            "warnung": "; ".join(warnungen) or None,
        }
        eintrag["fingerprint"] = fingerprint(eintrag)
        ergebnis.append(eintrag)
        statistik["gelesen"] += 1

    if not ergebnis:
        raise ValueError("Keine auswertbaren Zeilen gefunden.")
    ergebnis.sort(key=lambda e: (e["datum"], e["start"] or ""))
    return ergebnis, statistik
