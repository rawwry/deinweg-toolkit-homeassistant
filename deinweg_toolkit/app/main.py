"""Dein Weg Toolkit

Zeitlisten aus xlsx/csv zusammenfuehren, pruefen und als Nachweis exportieren.
Eingang ausschliesslich ueber den Upload in der Weboberflaeche.
"""

from __future__ import annotations

import asyncio
import calendar
import datetime as dt
import hashlib
import io
import os
import random
import re
import secrets
import tempfile
from urllib.parse import urlencode

from fastapi import (FastAPI, File, Form, HTTPException, Query, Request,
                     UploadFile)
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from markupsafe import Markup, escape
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response

from . import db
from . import mail
from .parser import (dauer_aus_spanne, fingerprint, hhmm, lies_datei, norm,
                     parse_datum, parse_dauer, parse_zeit, NICHT_ABRECHENBAR)
from . import auth
from . import wiki as _wiki

BASIS = os.path.dirname(__file__)

APP_NAME = os.environ.get("APP_NAME", "Dein Weg Toolkit")
VERSION = "1.7"

# Änderungsprotokoll, chronologisch von alt nach neu. Die Seite dreht die
# Reihenfolge selbst. Bewusst hier im Code und nicht in einer Textdatei, damit
# es beim Austausch des app-Ordners automatisch mitkommt.
from .changelog import CHANGELOG  # noqa: E402

MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "20"))
SPRUCH_DATEI = os.environ.get("SPRUCH_DATEI", "/texte/quotes.txt")
IDEEN_DATEI = os.environ.get("IDEEN_DATEI", "/texte/ideen.txt")
STRINGS_DATEI = os.environ.get("STRINGS_DATEI", "/texte/strings.txt")
WIKI_PFAD = os.environ.get("WIKI_PFAD", "/wiki")
FILES_PFAD = os.environ.get("FILES_PFAD", "/files")
# Sekunden zwischen zwei Pruefungen auf faellige E-Mail-Erinnerungen.
# 0 = Wecker aus. Standard: einmal pro Stunde.
WECKER_INTERVALL = int(os.environ.get("WECKER_INTERVALL", "3600"))

ENDUNGEN = ("xlsx", "xlsm", "csv")

app = FastAPI(title=APP_NAME, docs_url=None, redoc_url=None)
app.mount("/static", StaticFiles(directory=os.path.join(BASIS, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASIS, "templates"))
templates.env.filters["hhmm"] = hhmm
templates.env.globals["APP_NAME"] = APP_NAME
templates.env.globals["VERSION"] = VERSION
templates.env.globals["t"] = lambda *a, **k: t(*a, **k)
templates.env.globals["fusstext"] = lambda: fusstext()


# --- Anmeldung ---------------------------------------------------------------
#
# Ersetzt das frühere gemeinsame APP_PASSWORT (HTTP-Basic-Auth ohne einzelne
# Konten) durch echte Benutzerkonten mit Login, Rollen und Bereichsrechten.
# Die eigentliche Middleware, Sitzungsverwaltung und Zugriffsprüfung steckt
# in auth.py; hier wird sie nur eingebunden.

SITZUNG_TAGE = int(os.environ.get("SITZUNG_TAGE", "30"))

auth.setup(templates, SITZUNG_TAGE)
app.add_middleware(auth.SessionAuth)
app.include_router(auth.router)


# --- Hilfsfunktionen --------------------------------------------------------

def jetzt() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def deutsch(datum: str) -> str:
    try:
        return dt.date.fromisoformat(datum).strftime("%d.%m.%Y")
    except ValueError:
        return datum


templates.env.filters["deutsch"] = deutsch


# --- Texte für die Oberfläche ------------------------------------------------
#
# Alle erklärenden Texte stehen in strings.txt im Stammverzeichnis und können
# dort geändert werden, ohne den Code anzufassen. Fehlt ein Schlüssel oder die
# ganze Datei, greift der Standardtext von hier. Die Datei wird beim Start
# angelegt, falls sie noch nicht existiert.

from .texte_standard import TEXTE_STANDARD  # noqa: E402

_texte_zwischenspeicher: dict = {"stand": None, "werte": {}}


def texte() -> dict:
    """Liest strings.txt, sobald sie sich geändert hat."""
    try:
        stand = os.path.getmtime(STRINGS_DATEI)
    except OSError:
        stand = None

    if stand != _texte_zwischenspeicher["stand"]:
        werte = dict(TEXTE_STANDARD)
        if stand is not None:
            schluessel = None
            teile: list[str] = []
            try:
                with open(STRINGS_DATEI, encoding="utf-8") as f:
                    for zeile in f:
                        roh = zeile.rstrip("\n")
                        if roh.startswith("#") and not schluessel:
                            continue
                        kopf = re.fullmatch(r"\[([\w.]+)\]\s*", roh)
                        if kopf:
                            if schluessel:
                                werte[schluessel] = " ".join(" ".join(teile).split())
                            schluessel, teile = kopf.group(1), []
                        elif schluessel is not None:
                            teile.append(roh)
                if schluessel:
                    werte[schluessel] = " ".join(" ".join(teile).split())
            except OSError:
                pass
        _texte_zwischenspeicher.update({"stand": stand, "werte": werte})
    return _texte_zwischenspeicher["werte"]


def t(schluessel: str, **platzhalter) -> str:
    """Gibt den Text zurück, mit eingesetzten Platzhaltern."""
    text = texte().get(schluessel, "")
    if platzhalter:
        # Der Text selbst darf HTML enthalten (Links, <strong> in
        # strings.txt) und wird deshalb als Markup zurueckgegeben. Die
        # eingesetzten Werte stammen dagegen aus der Datenbank - etwa ein
        # Mitarbeitername - und werden vorher entschaerft, damit dort
        # kein HTML einschleusbar ist.
        try:
            text = text.format(**{k: escape(str(v))
                                  for k, v in platzhalter.items()})
        except (KeyError, IndexError, ValueError):
            pass
    return Markup(text)


def fusstext() -> str:
    """Der Fusszeilentext, mit dem Changelog-Link hinter der Version.

    Der Link steht bewusst hier und nicht in ``footer.text``: eine schon
    vorhandene ``strings.txt`` gewinnt gegen die Standardtexte, der Link
    waere dort also bei jeder bestehenden Installation unsichtbar
    geblieben. So haengt er am Markup und ist immer da.
    """
    roh = str(t("footer.text", version=VERSION))
    link = ' (<a href="/changelog">Changelog</a>)'
    # Erst hinter dem schliessenden </span> ansetzen: sonst stuende der
    # Link innerhalb von .version und waere mitgedaempft. Faellt die
    # Auszeichnung in einer eigenen strings.txt weg, greift der Rueckfall.
    if VERSION + "</span>" in roh:
        return Markup(roh.replace(VERSION + "</span>",
                                  VERSION + "</span>" + link, 1))
    return Markup(roh.replace(VERSION, VERSION + link, 1))


def strings_anlegen() -> None:
    """Schreibt strings.txt mit allen Standardtexten, falls sie fehlt."""
    if os.path.exists(STRINGS_DATEI):
        return
    zeilen = [
        "# Texte der Oberfläche. Änderungen wirken sofort, ohne Neustart.",
        "# Aufbau: [schluessel] in eckigen Klammern, darunter der Text.",
        "# Ein Text darf über mehrere Zeilen gehen, Umbrüche werden zu Leerzeichen.",
        "# Einfaches HTML wie <strong> oder <a href=\"...\"> ist erlaubt.",
        "# Geschweifte Klammern wie {zeitraum} sind Platzhalter und bleiben stehen.",
        "# Wird ein Schlüssel gelöscht, greift wieder der eingebaute Standardtext.",
        "",
    ]
    for schluessel, text in TEXTE_STANDARD.items():
        zeilen.append(f"[{schluessel}]")
        zeilen.append(text)
        zeilen.append("")
    try:
        os.makedirs(os.path.dirname(STRINGS_DATEI) or ".", exist_ok=True)
        with open(STRINGS_DATEI, "w", encoding="utf-8") as f:
            f.write("\n".join(zeilen))
        print(f"[start] {STRINGS_DATEI} angelegt", flush=True)
    except OSError as e:
        print(f"[start] strings.txt nicht schreibbar: {e}", flush=True)


def spruch() -> dict:
    """Zufälliger Block aus quotes.txt, getrennt in Zitat und Quelle.

    Die Blöcke sind durch eine Zeile mit ## voneinander getrennt. Beginnt die
    letzte Zeile eines Blocks mit einem Gedankenstrich, gilt sie als Quelle und
    wird kleiner gesetzt. Fehlt die Datei oder ist sie leer, kommt ein leerer
    Satz zurück und die Zeile auf der Startseite entfällt.
    """
    leer = {"text": "", "quelle": ""}
    try:
        with open(SPRUCH_DATEI, encoding="utf-8") as f:
            roh = f.read()
    except OSError:
        return leer

    bloecke = [b.strip("\n").strip() for b in re.split(r"^[ \t]*##[ \t]*$",
                                                       roh, flags=re.MULTILINE)]
    bloecke = [b for b in bloecke if b]
    if not bloecke:
        return leer

    zeilen = random.choice(bloecke).splitlines()
    quelle = ""
    if len(zeilen) > 1 and zeilen[-1].lstrip().startswith(("–", "—", "-", "~")):
        quelle = zeilen.pop().lstrip("–—-~ ").strip()
    return {"text": "\n".join(zeilen).strip(), "quelle": quelle}


def euro(betrag) -> str:
    """1234.5 wird zu 1.234,50 €"""
    try:
        text = f"{float(betrag):,.2f}"
    except (TypeError, ValueError):
        return "–"
    return text.replace(",", "#").replace(".", ",").replace("#", ".") + " €"


templates.env.filters["euro"] = euro


def stunden(wert) -> str:
    """7.5 wird zu '7,5', 4.0 zu '4' – für Eingabefelder und Anzeigen.

    Bewusst ohne Einheit: der Wert steht mal in einem Feld, mal im Text.
    Eine leere Ausgabe bei 0 waere hier falsch - in einem Eingabefeld soll
    die Null sichtbar sein.
    """
    try:
        zahl = float(wert or 0)
    except (TypeError, ValueError):
        return "0"
    text = f"{zahl:.2f}".rstrip("0").rstrip(".")
    return (text or "0").replace(".", ",")


templates.env.filters["stunden"] = stunden


def gesamtstunden(minuten) -> str:
    """5341 Minuten wird zu '89 Std 1 Min' – lesbar für grosse Summen.

    Im Unterschied zu hhmm (HH:MM, gedacht fuer einzelne Einheiten) ist das
    hier fuer Gesamtsummen wie den Bestand auf der Startseite gedacht, wo
    'HH:MM' bei mehreren tausend Stunden wie eine kaputte Uhrzeit aussieht.
    """
    try:
        minuten = int(minuten)
    except (TypeError, ValueError):
        return "0 Std"
    vorz = "-" if minuten < 0 else ""
    minuten = abs(minuten)
    stunden, rest = divmod(minuten, 60)
    text = f"{vorz}{stunden:,}".replace(",", ".") + " Std"
    if rest:
        text += f" {rest} Min"
    return text


templates.env.filters["gesamtstunden"] = gesamtstunden


def tage(wert) -> str:
    """2.0 wird zu '2 Tage', 1.0 zu '1 Tag', 2.5 bleibt '2,5 Tage'."""
    try:
        zahl = float(wert or 0)
    except (TypeError, ValueError):
        return str(wert)
    text = f"{zahl:.1f}".rstrip("0").rstrip(".").replace(".", ",")
    return f"{text} {'Tag' if abs(zahl) == 1 else 'Tage'}"


templates.env.filters["tage"] = tage


def sicherer_name(name: str) -> str:
    name = os.path.basename(name).replace("/", "_").replace("\\", "_")
    return re.sub(r"[^A-Za-z0-9._ äöüÄÖÜß-]", "_", name)[:120] or "datei"


# --- Kern: Datei einlesen ---------------------------------------------------

def verarbeite(dateiname: str, inhalt: bytes, mitarbeiter: str = "",
               erzwingen: bool = False, quelle: str = "Upload") -> int:
    """Liest eine Datei und legt Import samt Vorschauzeilen an.

    Die Originaldatei wird bewusst nicht aufgehoben: die Zeilen stehen
    anschliessend in der Datenbank, die Dateikopie waere nur Ballast. In
    "quelldatei" bleibt lediglich ein Vermerk mit Pruefsumme stehen.
    """
    zeilen, statistik = lies_datei(dateiname, inhalt, mitarbeiter, erzwingen)
    quell_hash = hashlib.sha256(inhalt).hexdigest()

    with db.db() as con:
        cur = con.execute(
            "INSERT INTO import (dateiname, mitarbeiter, hochgeladen_am, status, "
            "zeilen_gesamt, quelle, notiz) "
            "VALUES (?,?,?,'vorschau',?,?,?)",
            (dateiname, zeilen[0]["mitarbeiter"], jetzt(), len(zeilen), quelle,
             "Spalten: " + ", ".join(
                 f"{k}={v}" for k, v in statistik["spalten"].items())))
        import_id = cur.lastrowid

        gesehen: set[str] = set()
        neu = dubl = 0
        for z in zeilen:
            dublette = None
            if z["fingerprint"] in gesehen:
                dublette = "in dieser Datei doppelt"
            elif con.execute("SELECT 1 FROM eintrag WHERE fingerprint=?",
                             (z["fingerprint"],)).fetchone():
                dublette = "bereits in der Datenbank"
            gesehen.add(z["fingerprint"])
            if dublette:
                dubl += 1
            else:
                neu += 1
            con.execute(
                "INSERT INTO vorschau (import_id, mitarbeiter, datum, monat, start, "
                "ende, klient, beschreibung, dauer_min, abrechenbar, fingerprint, "
                "dublette, warnung) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (import_id, z["mitarbeiter"], z["datum"], z["monat"], z["start"],
                 z["ende"], z["klient"], z["beschreibung"], z["dauer_min"],
                 z["abrechenbar"], z["fingerprint"], dublette, z["warnung"]))
        con.execute("UPDATE import SET zeilen_neu=?, zeilen_dubletten=? WHERE id=?",
                    (neu, dubl, import_id))
        con.execute(
            "INSERT OR REPLACE INTO quelldatei (hash, dateiname, quelle, "
            "verarbeitet_am, import_id) VALUES (?,?,?,?,?)",
            (quell_hash, dateiname, quelle, jetzt(), import_id))
    return import_id


def uebernehmen_intern(import_id: int, mit_dubletten: bool = False) -> int:
    with db.db() as con:
        imp = con.execute("SELECT * FROM import WHERE id=?", (import_id,)).fetchone()
        if not imp or imp["status"] != "vorschau":
            raise HTTPException(400, "Import bereits abgeschlossen oder unbekannt")
        bedingung = "" if mit_dubletten else " AND dublette IS NULL"
        zeilen = con.execute(
            f"SELECT * FROM vorschau WHERE import_id=?{bedingung}",
            (import_id,)).fetchall()
        for z in zeilen:
            con.execute(
                "INSERT INTO eintrag (import_id, mitarbeiter, datum, monat, start, "
                "ende, klient, beschreibung, dauer_min, abrechenbar, fingerprint, "
                "angelegt_am) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (import_id, z["mitarbeiter"], z["datum"], z["monat"], z["start"],
                 z["ende"], z["klient"], z["beschreibung"], z["dauer_min"],
                 z["abrechenbar"], z["fingerprint"], jetzt()))
        con.execute("DELETE FROM vorschau WHERE import_id=?", (import_id,))
        con.execute("UPDATE import SET status='uebernommen', zeilen_neu=? WHERE id=?",
                    (len(zeilen), import_id))
    return len(zeilen)


async def wecker_schleife() -> None:
    """Prueft regelmaessig, ob Erinnerungen faellig sind.

    Stuendlich statt einmal taeglich: so wird eine Frist auch dann bemerkt,
    wenn der Container ueber Nacht aus war. Doppelte Mails verhindert der
    Vermerk in der Tabelle "benachrichtigung", nicht das Intervall.
    """
    await asyncio.sleep(30)
    while True:
        try:
            zeilen = await asyncio.to_thread(mail.durchlauf)
            for zeile in zeilen:
                if zeile not in ("nichts zu tun", "E-Mail-Versand ist ausgeschaltet"):
                    print(f"[wecker] {zeile}", flush=True)
        except Exception as e:
            print(f"[wecker] {type(e).__name__}: {e}", flush=True)
        await asyncio.sleep(WECKER_INTERVALL)


@app.on_event("startup")
async def start() -> None:
    initialer_admin = db.init()
    if initialer_admin:
        print("[start] " + "=" * 60, flush=True)
        print("[start] Erster Administrator angelegt:", flush=True)
        print(f"[start]   Benutzername: {initialer_admin['benutzername']}", flush=True)
        if initialer_admin["generiert"]:
            print(f"[start]   Passwort:     {initialer_admin['passwort']}", flush=True)
            print("[start]   (zufällig erzeugt, bitte nach dem ersten Login "
                  "unter Einstellungen ändern)", flush=True)
        else:
            print("[start]   Passwort:     wie in ADMIN_PASSWORT hinterlegt", flush=True)
        print("[start] " + "=" * 60, flush=True)
    strings_anlegen()
    _wiki.wiki_anlegen()
    try:
        os.makedirs(FILES_PFAD, exist_ok=True)
    except OSError as e:
        print(f"[start] {FILES_PFAD} nicht verfuegbar: {e}", flush=True)
    if WECKER_INTERVALL > 0:
        asyncio.create_task(wecker_schleife())
    print(f"[start] {APP_NAME} {VERSION} bereit · E-Mail-Wecker "
          f"{'alle ' + str(WECKER_INTERVALL) + 's' if WECKER_INTERVALL else 'aus'}",
          flush=True)


# --- Startseite -------------------------------------------------------------

def fehlerseite(text: str):
    return RedirectResponse("/?" + urlencode({"fehler": text}), status_code=303)


def zurueck_mit_hinweis(ziel: str, text: str):
    """Haengt einen Hinweis an eine Rueckkehradresse, die schon Filter traegt."""
    trenner = "&" if "?" in ziel else "?"
    return RedirectResponse(f"{ziel}{trenner}" + urlencode({"hinweis": text}),
                            status_code=303)


def abgabe_uebersicht(monat: str) -> dict:
    """Wer hat für diesen Monat schon Zeiten abgegeben und wer nicht.

    Abgegeben heißt: es liegen Datensätze für den Monat vor, egal ob importiert
    oder von Hand getippt. Der Abgleich läuft über die normalisierte Schreibweise
    des Namens, damit "timo" und "Timo " nicht auseinanderfallen.
    """
    with db.db() as con:
        team = con.execute(
            "SELECT * FROM mitarbeiter WHERE aktiv=1 "
            "ORDER BY abgabepflicht DESC, name").fetchall()
        zeilen = con.execute(
            "SELECT mitarbeiter, COUNT(*) n, SUM(dauer_min) m, "
            "MAX(angelegt_am) zuletzt, "
            "SUM(CASE WHEN import_id IS NULL THEN 1 ELSE 0 END) manuell "
            "FROM eintrag WHERE monat=? GROUP BY mitarbeiter", (monat,)).fetchall()

    nach_name = {norm(r["mitarbeiter"]): r for r in zeilen}
    zugeordnet = set()

    stand = []
    for m in team:
        treffer = nach_name.get(norm(m["name"]))
        if treffer:
            zugeordnet.add(norm(m["name"]))
        stand.append({
            "name": m["name"],
            "pflicht": bool(m["abgabepflicht"]),
            "da": bool(treffer),
            "n": treffer["n"] if treffer else 0,
            "m": treffer["m"] if treffer else 0,
            "zuletzt": treffer["zuletzt"] if treffer else "",
            "art": ("von Hand" if treffer and treffer["manuell"] == treffer["n"]
                    else ("importiert" if treffer and not treffer["manuell"]
                          else "gemischt" if treffer else "")),
        })

    # Namen, die in den Daten stehen, aber zu keinem Teammitglied passen
    fremd = [r["mitarbeiter"] for r in zeilen
             if norm(r["mitarbeiter"]) not in zugeordnet]

    pflichtige = [s for s in stand if s["pflicht"]]
    fertig = [s for s in pflichtige if s["da"]]
    return {
        "monat": monat,
        "wort": monat_wort(monat),
        "stand": stand,
        "fremd": sorted(fremd),
        "soll": len(pflichtige),
        "ist": len(fertig),
        "fehlend": [s["name"] for s in pflichtige if not s["da"]],
        "prozent": round(len(fertig) / len(pflichtige) * 100) if pflichtige else 0,
        "summe": sum(s["m"] for s in stand),
    }


def monat_verschieben(monat: str, schritte: int) -> str:
    jahr, mon = int(monat[:4]), int(monat[5:7])
    gesamt = jahr * 12 + (mon - 1) + schritte
    return f"{gesamt // 12:04d}-{gesamt % 12 + 1:02d}"


@app.get("/", response_class=HTMLResponse)
def startseite(request: Request, fehler: str = "", hinweis: str = "",
               alle: str = "", monat: str = "", mitarbeiter: str = "",
               datum: str = ""):
    """Zeiterfassung: manueller Eintrag und Listenimport auf einer Seite.

    Bis 2.6 waren das zwei getrennte Menuepunkte. Beide Wege fuehren zum
    selben Ergebnis - ein Datensatz in "eintrag" -, deshalb stehen sie
    jetzt untereinander auf einer Seite. Die Kaesten "Bestand" und
    "Abgaben" daneben bleiben unveraendert.
    """
    heute = dt.date.today()
    if not re.fullmatch(r"\d{4}-\d{2}", monat or ""):
        monat = heute.strftime("%Y-%m")
    mitarbeiter = mitarbeiter.strip()
    with db.db() as con:
        # Nur noch die Importe, die auf eine Pruefung warten. Die frueher
        # hier gezeigte Liste aller vergangenen Importe ist entfallen - die
        # Daten stehen in der Datenbank, die Dateiliste war nur Ballast.
        # Diese Auswahl bleibt aber noetig, sonst fuehrt kein Weg mehr zu
        # einer offenen Vorschau.
        importe = con.execute(
            "SELECT i.*, (SELECT COUNT(*) FROM vorschau v WHERE v.import_id=i.id) "
            "AS zeilen FROM import i WHERE i.status='vorschau' "
            "ORDER BY i.id DESC").fetchall()
        summe = con.execute(
            "SELECT COUNT(*) n, COALESCE(SUM(dauer_min),0) m FROM eintrag").fetchone()
        leute = [r["mitarbeiter"] for r in con.execute(
            "SELECT DISTINCT mitarbeiter FROM eintrag ORDER BY 1")]
        klienten = [r["klient"] for r in con.execute(
            "SELECT DISTINCT klient FROM eintrag ORDER BY 1")]
        leistungen = [r["name"] for r in con.execute(
            "SELECT name FROM leistung WHERE aktiv=1 ORDER BY name COLLATE NOCASE")]
        mitarbeiterliste = mitarbeiterauswahl(con)
        offene = con.execute(
            "SELECT COUNT(*) c FROM import WHERE status='vorschau'").fetchone()["c"]
        if mitarbeiter:
            letzte = con.execute(
                "SELECT * FROM eintrag WHERE mitarbeiter=? AND import_id IS NULL "
                "ORDER BY id DESC LIMIT 12", (mitarbeiter,)).fetchall()
            tag = parse_datum(datum)
            tagessumme = con.execute(
                "SELECT COALESCE(SUM(dauer_min),0) m, COUNT(*) n FROM eintrag "
                "WHERE mitarbeiter=? AND datum=?",
                (mitarbeiter, tag.isoformat() if tag else "")).fetchone()
        else:
            letzte, tagessumme = [], {"m": 0, "n": 0}
    return templates.TemplateResponse(request=request, name="index.html", context={
        "importe": importe, "summe": summe, "leute": leute,
        "klienten": klienten, "leistungen": leistungen, "letzte": letzte,
        "mitarbeiterliste": mitarbeiterliste,
        "tagessumme": tagessumme, "mitarbeiter": mitarbeiter, "datum": datum,
        "fehler": fehler, "hinweis": hinweis, "seite": "zeiterfassung",
        "offene": offene, "spruch": spruch(),
        "alle": bool(alle),
        "abgabe": abgabe_uebersicht(monat),
        "monat_vor": monat_verschieben(monat, -1),
        "monat_zurueck": monat_verschieben(monat, 1),
        "ist_laufender": monat == heute.strftime("%Y-%m")})


@app.post("/upload")
async def hochladen(datei: list[UploadFile] = File(...),
                    mitarbeiter: str = Form(""),
                    erzwingen: str = Form("")):
    letzte_id = None
    for f in datei:
        if not f.filename:
            continue
        inhalt = await f.read()
        if len(inhalt) > MAX_UPLOAD_MB * 1024 * 1024:
            return fehlerseite(f"{f.filename} ist größer als {MAX_UPLOAD_MB} MB.")
        try:
            letzte_id = verarbeite(f.filename, inhalt, mitarbeiter,
                                   bool(erzwingen), "Upload")
        except Exception as e:
            return fehlerseite(f"{f.filename}: {e}")
    if letzte_id is None:
        return fehlerseite("Keine Datei ausgewählt.")
    return RedirectResponse(f"/vorschau/{letzte_id}", status_code=303)


# --- Vorschau ---------------------------------------------------------------

@app.get("/vorschau/{import_id}", response_class=HTMLResponse)
def vorschau(request: Request, import_id: int):
    with db.db() as con:
        imp = con.execute("SELECT * FROM import WHERE id=?", (import_id,)).fetchone()
        if not imp:
            raise HTTPException(404, "Import nicht gefunden")
        zeilen = con.execute(
            "SELECT * FROM vorschau WHERE import_id=? ORDER BY datum, start",
            (import_id,)).fetchall()
        offene = con.execute(
            "SELECT id, dateiname FROM import WHERE status='vorschau' AND id<>?",
            (import_id,)).fetchall()
    return templates.TemplateResponse(request=request, name="vorschau.html", context={
        "imp": imp, "zeilen": zeilen, "offene": offene, "seite": "zeiterfassung",
        "summe_neu": sum(z["dauer_min"] for z in zeilen if not z["dublette"])})


@app.post("/vorschau/{import_id}/uebernehmen")
def uebernehmen(import_id: int, dubletten: str = Form("")):
    uebernehmen_intern(import_id, bool(dubletten))
    return RedirectResponse(f"/eintraege?import_id={import_id}", status_code=303)


@app.post("/vorschau/{import_id}/verwerfen")
def verwerfen(import_id: int):
    with db.db() as con:
        con.execute("DELETE FROM vorschau WHERE import_id=?", (import_id,))
        con.execute("DELETE FROM import WHERE id=? AND status='vorschau'", (import_id,))
    return RedirectResponse("/?" + urlencode(
        {"hinweis": "Import verworfen. Die Originaldatei bleibt im Archiv."}),
        status_code=303)


@app.post("/import/{import_id}/ruecknahme")
def ruecknahme(import_id: int):
    with db.db() as con:
        anzahl = con.execute("SELECT COUNT(*) c FROM eintrag WHERE import_id=?",
                             (import_id,)).fetchone()["c"]
        con.execute("DELETE FROM eintrag WHERE import_id=?", (import_id,))
        con.execute("UPDATE import SET status='zurueckgenommen' WHERE id=?",
                    (import_id,))
    return RedirectResponse("/?" + urlencode(
        {"hinweis": f"{anzahl} Einträge wieder entfernt."}), status_code=303)


# --- Eintraege --------------------------------------------------------------

MONATSNAMEN = {
    "01": "Januar", "02": "Februar", "03": "März", "04": "April",
    "05": "Mai", "06": "Juni", "07": "Juli", "08": "August",
    "09": "September", "10": "Oktober", "11": "November", "12": "Dezember",
}


def monat_wort(monat: str) -> str:
    """2026-08 wird zu 'August 2026'."""
    if not monat or len(monat) < 7:
        return monat or ""
    return f"{MONATSNAMEN.get(monat[5:7], monat[5:7])} {monat[:4]}"


templates.env.globals["monat_wort"] = monat_wort


def monatsliste(von: str, bis: str) -> list[str]:
    """Alle Monate von bis einschliesslich, als YYYY-MM."""
    try:
        jahr, mon = int(von[:4]), int(von[5:7])
        jahr_b, mon_b = int(bis[:4]), int(bis[5:7])
    except (ValueError, IndexError):
        return []
    ergebnis = []
    while (jahr, mon) <= (jahr_b, mon_b) and len(ergebnis) < 240:
        ergebnis.append(f"{jahr:04d}-{mon:02d}")
        mon += 1
        if mon > 12:
            mon, jahr = 1, jahr + 1
    return ergebnis


MONATSFAKTOR = 4.33  # durchschnittliche Wochen pro Monat, so von der Leitung vorgegeben


def runde_viertelstunde(minuten: float) -> int:
    """Rundet auf den nächsten 15-Minuten-Takt, damit keine krummen Werte entstehen."""
    return int(round(minuten / 15) * 15)


def soll_minuten(wochenstunden: float, monat: str = "") -> int | None:
    """Rechnet das Wochenkontingent mit dem Faktor 4,33 auf einen Monat hoch.

    Der Monat selbst geht nicht mehr in die Rechnung ein - das Ergebnis ist
    für jeden Monat gleich - wird aber als Parameter beibehalten, damit
    Aufrufer nicht angepasst werden müssen und ein leerer Monat weiterhin
    "kein Soll" bedeutet.
    """
    if not wochenstunden or not monat:
        return None
    return runde_viertelstunde(wochenstunden * 60 * MONATSFAKTOR)


def soll_zeitraum(wochenstunden: float, monate: list[str]) -> int | None:
    """Summiert das Soll über mehrere Monate."""
    if not wochenstunden or not monate:
        return None
    einzelsoll = soll_minuten(wochenstunden, monate[0])
    return einzelsoll * len(monate) if einzelsoll else None


# --- Bewilligte Zeiträume je betreuter Person --------------------------------
#
# Der Kostentraeger sagt Wochenstunden und Stundensatz immer nur befristet
# zu. "Michael Mueller" hat von 08/2024 bis 07/2025 vier Wochenstunden zu
# 65 EUR, ab 08/2025 sieben zu 70 EUR. Eine Auswertung ueber beide
# Zeitraeume muss deshalb Monat fuer Monat mit den Werten rechnen, die in
# diesem Monat galten - ein einziger Wert fuer den ganzen Zeitraum waere
# schlicht falsch.
#
# ⚠️ Gerechnet wird MONATSWEISE, obwohl die Zeitraeume taggenau erfasst
# werden. Ein Zeitraum gilt fuer jeden Monat, den er beruehrt. Alles andere
# waere Scheingenauigkeit: schon das Soll entsteht aus einem pauschalen
# Faktor 4,33 Wochen je Monat, und die Auswertung kennt ohnehin nur Monate.

def monatsgrenzen(monat: str) -> tuple[str, str]:
    """Erster und letzter Tag eines Monats als YYYY-MM-DD."""
    jahr, mon = int(monat[:4]), int(monat[5:7])
    letzter = (dt.date(jahr + (mon == 12), (mon % 12) + 1, 1)
               - dt.timedelta(days=1))
    return f"{monat}-01", letzter.isoformat()


# Ab wie vielen Tagen vor dem Ende gilt eine Bewilligung als "laeuft
# aus"? Zwei Monate - so lange dauert ein Folgeantrag beim Kostentraeger
# erfahrungsgemaess, und frueher waere es nur Rauschen.
BEWILLIGUNG_BALD_TAGE = 60


def bewilligungslage(zeitraeume, grund_stunden, grund_satz, heute: str) -> dict:
    """Wie steht eine betreute Person heute da?

    Eine Stelle fuer die Frage, die in den Einstellungen und in "Mein
    Bereich" gleich beantwortet werden muss. Die Liste kommt absteigend
    nach ``von`` herein, wie ueberall (siehe kontingent_im_monat).

    ``art`` ist eines von:
    * ``laufend``    - alles in Ordnung
    * ``laeuft_aus`` - gilt noch, endet aber in den naechsten 60 Tagen
    * ``abgelaufen`` - der letzte Bescheid ist vorbei
    * ``kuenftig``   - der naechste beginnt erst
    * ``grundwert``  - gar kein Bescheid, aber ein Grundwert
    * ``leer``       - nichts hinterlegt
    """
    laufend = None
    for z in zeitraeume or []:
        if z["von"] <= heute and (not z["bis"] or z["bis"] >= heute):
            laufend = z
            break

    if laufend:
        if laufend["bis"]:
            try:
                tage = (dt.date.fromisoformat(laufend["bis"])
                        - dt.date.fromisoformat(heute)).days
            except ValueError:
                tage = None
            if tage is not None and tage <= BEWILLIGUNG_BALD_TAGE:
                return {"art": "laeuft_aus", "bis": laufend["bis"],
                        "tage": tage, "zeitraum": laufend}
        return {"art": "laufend", "zeitraum": laufend}

    if zeitraeume:
        kuenftig = [z for z in zeitraeume if z["von"] > heute]
        if kuenftig:
            naechster = min(kuenftig, key=lambda z: z["von"])
            return {"art": "kuenftig", "ab": naechster["von"],
                    "zeitraum": naechster}
        vergangen = [z for z in zeitraeume if z["bis"] and z["bis"] < heute]
        letzter = max(vergangen, key=lambda z: z["bis"]) if vergangen else None
        return {"art": "abgelaufen",
                "seit": letzter["bis"] if letzter else "", "zeitraum": letzter}

    if grund_stunden or grund_satz:
        return {"art": "grundwert"}
    return {"art": "leer"}


# Welche Lagen verlangen, dass jemand tätig wird? Reihenfolge ist zugleich
# die Dringlichkeit, nach der sortiert wird.
BEWILLIGUNG_HANDLUNG = ("abgelaufen", "leer", "laeuft_aus", "kuenftig",
                        "grundwert")


def bewilligungen_pruefen(con) -> list[dict]:
    """Alle aktiven betreuten Personen, bei denen etwas zu tun ist.

    Dringendstes zuerst. Grundlage fuer die Karte in "Mein Bereich".

    ⚠️ Die Karte trennt danach in zwei Gruppen: alles ausser "grundwert"
    verlangt einen Folgeantrag, "grundwert" ist nur eine Feststellung
    ("hier war nie ein Bescheid hinterlegt"). Ohne diese Trennung
    ertraenken zwanzig Grundwert-Zeilen die drei, um die es geht.
    """
    heute = dt.date.today().isoformat()
    zr = zeitraeume_lesen(con)
    offen = []
    for p in con.execute(
            "SELECT id, name, wochenstunden, stundensatz FROM person "
            "WHERE aktiv=1 ORDER BY name"):
        stand = bewilligungslage(zr.get(p["name"], []), p["wochenstunden"],
                                 p["stundensatz"], heute)
        if stand["art"] in BEWILLIGUNG_HANDLUNG:
            offen.append({**stand, "name": p["name"], "id": p["id"],
                          "rang": BEWILLIGUNG_HANDLUNG.index(stand["art"])})
    offen.sort(key=lambda r: (r["rang"], r["name"]))
    return offen


def zeitraeume_lesen(con) -> dict[str, list]:
    """Alle bewilligten Zeiträume, nach Personennamen gebündelt.

    Sortiert nach ``von`` absteigend: der zuletzt begonnene Zeitraum steht
    vorn und gewinnt damit bei Überschneidungen (siehe kontingent_im_monat).
    """
    je_person: dict[str, list] = {}
    for r in con.execute(
            "SELECT p.name, z.* FROM person_zeitraum z "
            "JOIN person p ON p.id = z.person_id "
            "ORDER BY z.von DESC, z.id DESC"):
        je_person.setdefault(r["name"], []).append(r)
    return je_person


def kontingent_im_monat(monat: str, zeitraeume, grund_stunden: float,
                        grund_satz: float) -> tuple[float, float, bool]:
    """Welche Wochenstunden und welcher Stundensatz galten in diesem Monat?

    Gibt ``(wochenstunden, stundensatz, aus_zeitraum)`` zurück. Greift kein
    Zeitraum, gelten die Grundwerte der Person — so rechnen alle bisher
    gepflegten Personen unverändert weiter.

    ⚠️ **Überschneiden sich zwei Zeiträume, gewinnt der später begonnene.**
    Das kommt in der Praxis vor, wenn ein Folgebescheid schon läuft,
    während der alte formal noch nicht abgelaufen ist. Die Liste kommt
    absteigend nach ``von`` herein, der erste Treffer ist also der
    richtige.
    """
    anfang, ende = monatsgrenzen(monat)
    for z in zeitraeume or []:
        if z["von"] > ende:
            continue
        if z["bis"] and z["bis"] < anfang:
            continue
        return (z["wochenstunden"] or 0), (z["stundensatz"] or 0), True
    return grund_stunden or 0, grund_satz or 0, False


def bereichsfilter(von_jahr="", von_monat="", bis_jahr="", bis_monat="",
                   mitarbeiter="", klient="", q="", import_id=0,
                   nur_abrechenbar=""):
    """Ein Filter für Auswertung, Datensätze und Export – bewusst nur einmal.

    Baut aus den Formularfeldern die WHERE-Bedingung, die Beschriftung des
    Zeitraums und die Parameter für Links zurück.
    """
    von_jahr, von_monat = str(von_jahr or "").strip(), str(von_monat or "").strip()
    bis_jahr, bis_monat = str(bis_jahr or "").strip(), str(bis_monat or "").strip()

    # "klient" und "mitarbeiter" nehmen seit 1.5 bzw. 1.6 mehrere Namen
    # entgegen - man wertet oft zwei oder drei zusammen aus. Ein einzelner
    # String kommt weiterhin an (alte Lesezeichen, Links aus anderen
    # Seiten) und wird hier zur Liste mit einem Element.
    def als_liste(wert) -> list[str]:
        if isinstance(wert, str):
            roh = [wert.strip()] if wert.strip() else []
        else:
            roh = [str(w).strip() for w in (wert or []) if str(w).strip()]
        # Reihenfolge stabil halten, Dubletten raus - sonst steht derselbe
        # Name zweimal in der Chipleiste.
        return list(dict.fromkeys(roh))

    klienten_filter = als_liste(klient)
    leute_filter = als_liste(mitarbeiter)

    # Teilangaben sinnvoll ergänzen: ein Jahr ohne Monat meint das ganze Jahr.
    von = f"{von_jahr}-{von_monat or '01'}" if von_jahr else ""
    bis = f"{bis_jahr}-{bis_monat or '12'}" if bis_jahr else ""
    if von and bis and von > bis:
        von, bis = bis, von
        von_jahr, von_monat, bis_jahr, bis_monat = (
            bis_jahr, bis_monat, von_jahr, von_monat)

    # Ein Monat ohne Jahr meint diesen Monat in allen Jahren
    nur_monate = (von_monat if not von_jahr else "", bis_monat if not bis_jahr else "")

    wo, werte = ["1=1"], []
    if von:
        wo.append("monat>=?"); werte.append(von)
    if bis:
        wo.append("monat<=?"); werte.append(bis)
    if nur_monate[0] and nur_monate[1]:
        a, b = sorted(nur_monate)
        wo.append("substr(monat, 6, 2) BETWEEN ? AND ?"); werte += [a, b]
    elif nur_monate[0]:
        wo.append("substr(monat, 6, 2)>=?"); werte.append(nur_monate[0])
    elif nur_monate[1]:
        wo.append("substr(monat, 6, 2)<=?"); werte.append(nur_monate[1])
    if leute_filter:
        wo.append("mitarbeiter IN (" + ",".join("?" * len(leute_filter)) + ")")
        werte += leute_filter
    if klienten_filter:
        platzhalter = ",".join("?" * len(klienten_filter))
        wo.append(f"klient IN ({platzhalter})")
        werte += klienten_filter
    if import_id:
        wo.append("import_id=?"); werte.append(import_id)
    if nur_abrechenbar:
        wo.append("klient IN (SELECT name FROM person WHERE abrechenbar=1)")
    if q:
        wo.append("(beschreibung LIKE ? OR klient LIKE ? OR mitarbeiter LIKE ?)")
        werte += [f"%{q}%"] * 3

    def monatsname(nr):
        return MONATSNAMEN.get(nr, nr)

    if von and bis and von == bis:
        wort = monat_wort(von)
    elif von and bis and von[:4] == bis[:4] and von[5:] == "01" and bis[5:] == "12":
        wort = f"Jahr {von[:4]}"
    elif von and bis:
        wort = f"{monat_wort(von)} bis {monat_wort(bis)}"
    elif von:
        wort = f"ab {monat_wort(von)}"
    elif bis:
        wort = f"bis {monat_wort(bis)}"
    elif nur_monate[0] and nur_monate[1] and nur_monate[0] == nur_monate[1]:
        wort = f"{monatsname(nur_monate[0])}, alle Jahre"
    elif nur_monate[0] and nur_monate[1]:
        a, b = sorted(nur_monate)
        wort = f"{monatsname(a)} bis {monatsname(b)}, alle Jahre"
    elif nur_monate[0]:
        wort = f"ab {monatsname(nur_monate[0])}, alle Jahre"
    elif nur_monate[1]:
        wort = f"bis {monatsname(nur_monate[1])}, alle Jahre"
    else:
        wort = "alle Zeiten"

    felder = {"von_jahr": von_jahr, "von_monat": von_monat,
              "bis_jahr": bis_jahr, "bis_monat": bis_monat,
              # Der Einzelwert bleibt (fuer Links und die alte
              # Schreibweise), daneben steht die vollstaendige Liste.
              "mitarbeiter": leute_filter[0] if len(leute_filter) == 1 else "",
              "mitarbeiterliste": leute_filter,
              "klient": klienten_filter[0] if len(klienten_filter) == 1 else "",
              "klienten": klienten_filter, "q": q,
              "import_id": import_id or "", "nur_abrechenbar": nur_abrechenbar}

    aktive = []
    if von or bis or nur_monate[0] or nur_monate[1]:
        aktive.append(("Zeitraum", wort))
    if len(klienten_filter) == 1:
        aktive.append(("Betreute Person", klienten_filter[0]))
    elif klienten_filter:
        aktive.append((f"{len(klienten_filter)} betreute Personen",
                       ", ".join(klienten_filter)))
    if len(leute_filter) == 1:
        aktive.append(("Mitarbeiter", leute_filter[0]))
    elif leute_filter:
        aktive.append((f"{len(leute_filter)} Mitarbeiter", ", ".join(leute_filter)))
    if q:
        aktive.append(("Suche", q))
    if nur_abrechenbar:
        aktive.append(("Nur abrechenbar", "ja"))
    if import_id:
        aktive.append(("Import", f"Nr. {import_id}"))

    return {
        "wo": " AND ".join(wo), "werte": werte,
        "von": von, "bis": bis, "wort": wort, "nur_monate": nur_monate,
        "f": felder, "aktive": aktive,
        # doseq, damit mehrere Namen als eigene klient=-Parameter
        # herauskommen und nicht als ein String mit Kommas.
        # doseq, damit mehrere Namen als eigene klient=/mitarbeiter=
        # Parameter herauskommen und nicht als ein String mit Kommas.
        "query": urlencode(
            {k: v for k, v in felder.items()
             if v and k not in ("klient", "klienten",
                                "mitarbeiter", "mitarbeiterliste")}
            | ({"klient": klienten_filter} if klienten_filter else {})
            | ({"mitarbeiter": leute_filter} if leute_filter else {}),
            doseq=True),
    }


def mitarbeiterauswahl(con) -> dict:
    """Namen fuer das Auswahlfeld "Mitarbeiter".

    Erste Gruppe ist das gepflegte Team (Einstellungen -> Mitarbeiter),
    nur die aktiven. Zweite Gruppe sind Namen, die in den Zeiten schon
    vorkommen, aber nicht (mehr) im Team stehen - sonst waeren Zeiten von
    ausgeschiedenen Kolleginnen ueber das Auswahlfeld nicht mehr
    erreichbar. Verglichen wird ueber norm(), weil Schreibweisen aus
    Fremdexporten abweichen koennen.
    """
    team = [r["name"] for r in con.execute(
        "SELECT name FROM mitarbeiter WHERE aktiv=1 ORDER BY name COLLATE NOCASE")]
    bekannt = {norm(n) for n in team}
    weitere = [r["mitarbeiter"] for r in con.execute(
        "SELECT DISTINCT mitarbeiter FROM eintrag ORDER BY 1 COLLATE NOCASE")
        if norm(r["mitarbeiter"]) not in bekannt]
    return {"team": team, "weitere": weitere}


def auswahllisten() -> dict:
    """Jahre, Mitarbeiter und betreute Personen für die Auswahlfelder."""
    heute = dt.date.today()
    with db.db() as con:
        jahre = [r["j"] for r in con.execute(
            "SELECT DISTINCT substr(monat, 1, 4) j FROM eintrag ORDER BY 1 DESC")]
        leute = [r["mitarbeiter"] for r in con.execute(
            "SELECT DISTINCT mitarbeiter FROM eintrag ORDER BY 1")]
        klienten = [r["klient"] for r in con.execute(
            "SELECT DISTINCT klient FROM eintrag ORDER BY 1")]
    if str(heute.year) not in jahre:
        jahre = [str(heute.year)] + jahre
    return {"jahre": jahre, "leute": leute, "klienten": klienten,
            "monatsnamen": MONATSNAMEN}


# --- Wer darf welchen Zeiteintrag loeschen? ---------------------------------
#
# Regel: die eigenen Eintraege darf jeder loeschen, immer. Fuer die
# Eintraege anderer braucht es das ausdrueckliche Recht "fremde_loeschen"
# (Einstellungen -> Benutzerverwaltung); Administratoren haben es ohnehin.
#
# "Eigen" heisst: eintrag.mitarbeiter entspricht dem Mitarbeiter, der zum
# angemeldeten Konto gehoert. Die Zuordnung laeuft ueber dieselbe Regel wie
# beim E-Mail-Versand und in "Mein Bereich" (mitarbeiter_zu_benutzer), der
# Vergleich ueber norm() - Schreibweisen aus Fremdexporten stimmen nicht
# immer aufs Zeichen ueberein.

def eigener_mitarbeitername(con, benutzer) -> str:
    """Der Mitarbeitername des angemeldeten Kontos, oder "" wenn keiner passt."""
    zeile = mitarbeiter_zu_benutzer(con, benutzer)
    if zeile is None:
        return ""
    try:
        return (zeile["name"] or "").strip()
    except (IndexError, KeyError, TypeError):
        return ""


def _ist_eigener(eintrag_mitarbeiter: str, eigener_name: str) -> bool:
    if not eigener_name:
        return False
    return norm(eintrag_mitarbeiter) == norm(eigener_name)


def darf_eintrag_loeschen(benutzer, eintrag_mitarbeiter: str,
                          eigener_name: str) -> bool:
    return (auth.darf_fremde_loeschen(benutzer)
            or _ist_eigener(eintrag_mitarbeiter, eigener_name))


def darf_eintrag_bearbeiten(benutzer, eintrag_mitarbeiter: str,
                            eigener_name: str) -> bool:
    return (auth.darf_fremde_bearbeiten(benutzer)
            or _ist_eigener(eintrag_mitarbeiter, eigener_name))


@app.get("/eintraege", response_class=HTMLResponse)
def eintraege(request: Request, von_jahr: str = "", von_monat: str = "",
              bis_jahr: str = "", bis_monat: str = "",
              mitarbeiter: list[str] = Query([]),
              klient: list[str] = Query([]), q: str = "", import_id: int = 0,
              nur_abrechenbar: str = "", seite_nr: int = 1, hinweis: str = ""):
    filter_ = bereichsfilter(von_jahr, von_monat, bis_jahr, bis_monat,
                             mitarbeiter, klient, q, import_id, nur_abrechenbar)
    wo, werte = filter_["wo"], filter_["werte"]
    pro_seite = 200

    with db.db() as con:
        kopf = con.execute(
            f"SELECT COUNT(*) n, COALESCE(SUM(dauer_min),0) m FROM eintrag WHERE {wo}",
            werte).fetchone()
        # Seitenzahl vor der Abfrage begrenzen, sonst kommt eine leere Seite
        seiten_gesamt = max(1, -(-kopf["n"] // pro_seite))
        seite_nr = min(max(1, seite_nr), seiten_gesamt)
        zeilen = con.execute(
            f"SELECT * FROM eintrag WHERE {wo} ORDER BY datum DESC, start DESC "
            f"LIMIT {pro_seite} OFFSET {(seite_nr - 1) * pro_seite}", werte).fetchall()
        eigener = eigener_mitarbeitername(con, request.state.benutzer)

    # Welche der angezeigten Zeilen darf dieses Konto loeschen? Einmal hier
    # gerechnet statt in der Vorlage - dieselbe Funktion entscheidet auch
    # serverseitig beim Loeschen, so koennen Ansicht und Durchsetzung nicht
    # auseinanderlaufen.
    darf_fremde = auth.darf_fremde_loeschen(request.state.benutzer)
    darf_fremde_bearb = auth.darf_fremde_bearbeiten(request.state.benutzer)
    loeschbar = {z["id"] for z in zeilen
                 if darf_eintrag_loeschen(request.state.benutzer,
                                          z["mitarbeiter"], eigener)}
    bearbeitbar = {z["id"] for z in zeilen
                   if darf_eintrag_bearbeiten(request.state.benutzer,
                                              z["mitarbeiter"], eigener)}

    zusatz = auswahllisten()
    return templates.TemplateResponse(request=request, name="eintraege.html", context={
        "zeilen": zeilen, "kopf": kopf, "f": filter_["f"], "seite_nr": seite_nr,
        "loeschbar": loeschbar, "bearbeitbar": bearbeitbar,
        "darf_fremde": darf_fremde, "darf_fremde_bearb": darf_fremde_bearb,
        "eigener": eigener,
        "hinweis": hinweis, "aktive_filter": filter_["aktive"],
        "zeitraum_wort": filter_["wort"], "query": filter_["query"],
        "mehr": kopf["n"] > seite_nr * pro_seite,
        "seiten_gesamt": seiten_gesamt, "pro_seite": pro_seite,
        "erste_nr": (seite_nr - 1) * pro_seite + 1,
        "letzte_nr": min(seite_nr * pro_seite, kopf["n"]),
        "seite": "eintraege", **zusatz})


@app.post("/eintraege/loeschen")
def eintraege_sammelloeschen(request: Request, ids: list[int] = Form([]),
                             zurueck: str = Form("/eintraege")):
    """Loescht mehrere angekreuzte Datensaetze auf einmal.

    Die Sicherheitsabfrage sitzt im Browser (siehe eintraege.html); hier
    wird geprueft, dass ueberhaupt etwas angekreuzt war - und dass jeder
    einzelne Datensatz auch geloescht werden darf. Die Kaestchen fehlender
    Rechte blendet die Vorlage zwar aus, ein von Hand abgeschicktes
    Formular kaeme sonst aber trotzdem durch. Die Obergrenze entspricht
    einer vollen Seite der Datensatzliste.
    """
    ids = ids[:500]
    if not ids:
        return zurueck_mit_hinweis(zurueck, "Es war nichts angekreuzt.")
    benutzer = request.state.benutzer
    with db.db() as con:
        eigener = eigener_mitarbeitername(con, benutzer)
        platzhalter = ",".join("?" * len(ids))
        vorhanden = con.execute(
            f"SELECT id, mitarbeiter FROM eintrag WHERE id IN ({platzhalter})",
            ids).fetchall()
        erlaubt = [z["id"] for z in vorhanden
                   if darf_eintrag_loeschen(benutzer, z["mitarbeiter"], eigener)]
        verweigert = len(vorhanden) - len(erlaubt)
        if erlaubt:
            platzhalter = ",".join("?" * len(erlaubt))
            con.execute(f"DELETE FROM eintrag WHERE id IN ({platzhalter})", erlaubt)

    if not erlaubt:
        return zurueck_mit_hinweis(
            zurueck, "Nichts gelöscht: Das waren ausschließlich Einträge "
                     "anderer Personen. Dafür fehlt dir die Berechtigung.")
    wort = "Eintrag" if len(erlaubt) == 1 else "Einträge"
    text = f"{len(erlaubt)} {wort} gelöscht."
    if verweigert:
        wort2 = "Eintrag" if verweigert == 1 else "Einträge"
        text += (f" {verweigert} {wort2} von anderen Personen "
                 "blieben stehen – dafür fehlt dir die Berechtigung.")
    return zurueck_mit_hinweis(zurueck, text)


# Dieselbe Route zusaetzlich unter /meinbereich: „Mein Bereich" haengt
# an keiner Bereichsberechtigung, jeder soll dort seine eigenen Zeiten
# aendern und loeschen koennen - auch ohne den Bereich „Übersicht
# (Datensätze)". Die Pruefung bleibt dieselbe: darf_eintrag_loeschen
# entscheidet ueber den Eintrag, nicht ueber den Pfad. Ein fremder
# Eintrag wird hier also genauso abgewiesen wie dort.
@app.post("/eintraege/{eintrag_id}/loeschen")
@app.post("/meinbereich/eintrag/{eintrag_id}/loeschen")
def eintrag_loeschen(request: Request, eintrag_id: int,
                     zurueck: str = Form("/eintraege")):
    benutzer = request.state.benutzer
    with db.db() as con:
        z = con.execute("SELECT mitarbeiter FROM eintrag WHERE id=?",
                        (eintrag_id,)).fetchone()
        if z is None:
            return zurueck_mit_hinweis(zurueck, "Diesen Eintrag gibt es nicht mehr.")
        eigener = eigener_mitarbeitername(con, benutzer)
        if not darf_eintrag_loeschen(benutzer, z["mitarbeiter"], eigener):
            return zurueck_mit_hinweis(
                zurueck, f"„{z['mitarbeiter']}“ ist nicht dein Eintrag. "
                         "Zum Löschen fremder Einträge fehlt dir die Berechtigung.")
        con.execute("DELETE FROM eintrag WHERE id=?", (eintrag_id,))
    return RedirectResponse(zurueck, status_code=303)


@app.get("/eintraege/{eintrag_id}/bearbeiten", response_class=HTMLResponse)
@app.get("/meinbereich/eintrag/{eintrag_id}/bearbeiten", response_class=HTMLResponse)
def eintrag_formular(request: Request, eintrag_id: int,
                     zurueck: str = "/eintraege", fehler: str = ""):
    with db.db() as con:
        z = con.execute("SELECT * FROM eintrag WHERE id=?", (eintrag_id,)).fetchone()
        if not z:
            raise HTTPException(404, "Eintrag nicht gefunden")
        # Gleiche Regel wie beim Loeschen: die eigenen immer, fremde nur mit
        # dem Recht. Schon hier und nicht erst beim Speichern - sonst tippt
        # man erst einen Text und bekommt danach die Abfuhr.
        if not darf_eintrag_bearbeiten(request.state.benutzer,
                                       z["mitarbeiter"],
                                       eigener_mitarbeitername(
                                           con, request.state.benutzer)):
            return zurueck_mit_hinweis(
                zurueck, f"„{z['mitarbeiter']}“ ist nicht dein Eintrag. "
                         "Zum Bearbeiten fremder Einträge fehlt dir die "
                         "Berechtigung.")
        leute = [r["mitarbeiter"] for r in con.execute(
            "SELECT DISTINCT mitarbeiter FROM eintrag ORDER BY 1")]
        klienten = [r["klient"] for r in con.execute(
            "SELECT DISTINCT klient FROM eintrag ORDER BY 1")]
    return templates.TemplateResponse(request=request, name="bearbeiten.html", context={
        "z": z, "leute": leute, "klienten": klienten, "zurueck": zurueck,
        "fehler": fehler,
        "seite": "meinbereich" if request.url.path.startswith("/meinbereich")
                 else "eintraege"})


@app.post("/eintraege/{eintrag_id}/bearbeiten")
@app.post("/meinbereich/eintrag/{eintrag_id}/bearbeiten")
def eintrag_speichern(request: Request, eintrag_id: int,
                      datum: str = Form(...), start: str = Form(""),
                      ende: str = Form(""), dauer: str = Form(""),
                      klient: str = Form(...), beschreibung: str = Form(""),
                      mitarbeiter: str = Form(...), abrechenbar: str = Form(""),
                      zurueck: str = Form("/eintraege")):

    def zurueck_mit_fehler(text: str):
        return RedirectResponse(
            request.url.path + "?" +
            urlencode({"zurueck": zurueck, "fehler": text}), status_code=303)

    try:
        tag = dt.date.fromisoformat(datum.strip())
    except ValueError:
        return zurueck_mit_fehler("Das Datum ist nicht gültig.")

    beginn = parse_zeit(start) if start.strip() else None
    schluss = parse_zeit(ende) if ende.strip() else None

    minuten = parse_dauer(dauer) if dauer.strip() else None
    if minuten is None:
        minuten = dauer_aus_spanne(beginn, schluss)
    if minuten is None:
        return zurueck_mit_fehler(
            "Die Dauer fehlt. Trag sie als Stunden:Minuten ein, "
            "zum Beispiel 01:30, oder gib Beginn und Ende an.")
    if minuten <= 0:
        return zurueck_mit_fehler("Die Dauer muss größer als null sein.")

    klient = klient.strip() or "Ohne Zuordnung"
    mitarbeiter = mitarbeiter.strip()
    if not mitarbeiter:
        return zurueck_mit_fehler("Ohne Mitarbeiter geht es nicht.")

    neu = {"mitarbeiter": mitarbeiter, "datum": tag.isoformat(), "start": beginn,
           "ende": schluss, "klient": klient,
           "beschreibung": beschreibung.strip(), "dauer_min": int(minuten)}

    with db.db() as con:
        # Gegen den Stand in der Datenbank pruefen, nicht gegen das Formular:
        # der Mitarbeitername steht als aenderbares Feld darin, sonst koennte
        # man ihn beim Speichern einfach auf den eigenen umbiegen.
        vorher = con.execute("SELECT mitarbeiter FROM eintrag WHERE id=?",
                             (eintrag_id,)).fetchone()
        if vorher is None:
            return zurueck_mit_hinweis(zurueck, "Diesen Eintrag gibt es nicht mehr.")
        eigener = eigener_mitarbeitername(con, request.state.benutzer)
        if not darf_eintrag_bearbeiten(request.state.benutzer,
                                       vorher["mitarbeiter"], eigener):
            return zurueck_mit_hinweis(
                zurueck, f"„{vorher['mitarbeiter']}“ ist nicht dein Eintrag. "
                         "Zum Bearbeiten fremder Einträge fehlt dir die "
                         "Berechtigung.")
        # Ebenso wenig darf man ihn auf jemand anderen umschreiben, wenn
        # man fremde Eintraege gar nicht anfassen duerfte.
        if not darf_eintrag_bearbeiten(request.state.benutzer, mitarbeiter,
                                       eigener):
            return zurueck_mit_hinweis(
                zurueck, "Du kannst den Eintrag nicht auf "
                         f"„{mitarbeiter}“ umschreiben – dafür fehlt dir "
                         "die Berechtigung.")
        con.execute(
            "UPDATE eintrag SET mitarbeiter=?, datum=?, monat=?, start=?, ende=?, "
            "klient=?, beschreibung=?, dauer_min=?, abrechenbar=?, fingerprint=? "
            "WHERE id=?",
            (mitarbeiter, neu["datum"], tag.strftime("%Y-%m"), beginn, schluss,
             klient, neu["beschreibung"], int(minuten),
             1 if abrechenbar else 0, fingerprint(neu), eintrag_id))

    trenner = "&" if "?" in zurueck else "?"
    return RedirectResponse(
        f"{zurueck}{trenner}" + urlencode({"hinweis": "Eintrag gespeichert."}),
        status_code=303)


# --- Manuelle Zeiterfassung -------------------------------------------------

def zeit_locker(text: str) -> str | None:
    """Nimmt 14:30, 14.30 und 1430 gleichermassen an."""
    text = (text or "").strip()
    if not text:
        return None
    if re.fullmatch(r"\d{3,4}", text):
        text = text.zfill(4)[:-2] + ":" + text.zfill(4)[-2:]
    return parse_zeit(text)


@app.get("/erfassung")
def erfassung_umleitung(mitarbeiter: str = "", datum: str = ""):
    """Die manuelle Erfassung wohnt seit 0.6.10 auf der Startseite.

    Die Adresse bleibt bestehen, damit gespeicherte Lesezeichen und die
    Rueckverweise aus dem Bearbeiten-Formular nicht ins Leere laufen.
    """
    werte = {k: v for k, v in (("mitarbeiter", mitarbeiter), ("datum", datum)) if v}
    return RedirectResponse("/" + ("?" + urlencode(werte) if werte else ""),
                            status_code=303)


@app.post("/erfassung")
def erfassung_speichern(mitarbeiter: str = Form(""),
                        datum: list[str] = Form([]),
                        klient: list[str] = Form([]),
                        start: list[str] = Form([]),
                        ende: list[str] = Form([]),
                        leistung: list[str] = Form([]),
                        beschreibung: list[str] = Form([])):
    """Legt einen oder mehrere Eintraege an.

    Die Felder kommen als parallele Listen an: jede Zeile des Formulars
    schickt genau einen Wert je Feld, in Dokumentreihenfolge - damit
    stehen die Listen zeilenweise untereinander. Bei nur einer Zeile ist
    das dieselbe Sache wie frueher, nur eben mit einem Element.

    Entweder alles oder nichts: Findet sich in irgendeiner Zeile ein
    Fehler, wird gar nichts gespeichert und die Meldung nennt die
    Zeilennummer. Halb gespeicherte Stapel waeren schlimmer als ein
    Abbruch - man wuesste hinterher nicht, was schon drin steht.
    """
    mitarbeiter = mitarbeiter.strip()
    erstes_datum = (datum[0].strip() if datum else "")

    def zurueck(**mehr):
        werte = {"mitarbeiter": mitarbeiter, "datum": erstes_datum}
        werte.update(mehr)
        # Die Sprungmarke haelt die Seite beim Formular stehen, statt nach
        # dem Speichern wieder ganz oben zu landen.
        return RedirectResponse("/?" + urlencode(werte) + "#erfassen",
                                status_code=303)

    if not mitarbeiter:
        return zurueck(fehler="Wähl oben aus, wer die Zeiten erfasst.")

    # Die Listen auf gleiche Laenge bringen. leistung fehlt ganz, solange
    # keine Leistungsbeschreibungen gepflegt sind - dann gibt es das
    # Auswahlfeld gar nicht.
    anzahl = max(len(datum), len(klient), len(start), len(ende))
    if not anzahl:
        return zurueck(fehler="Es war keine Zeile ausgefüllt.")

    def feld(liste: list[str], nr: int) -> str:
        return (liste[nr] if nr < len(liste) else "").strip()

    saetze = []
    for nr in range(anzahl):
        d, k = feld(datum, nr), feld(klient, nr)
        a, e = feld(start, nr), feld(ende, nr)
        l, b = feld(leistung, nr), feld(beschreibung, nr)
        zeile = nr + 1

        # Vollstaendig leere Zeilen einfach ueberspringen. Sie entstehen,
        # wenn jemand eine Zeile hinzufuegt und dann doch nicht braucht.
        if not any((d, k, a, e, l, b)):
            continue

        fehlt = [name for name, wert in
                 (("Datum", d), ("Betreuter", k), ("Startzeit", a), ("Endzeit", e))
                 if not wert]
        if fehlt:
            return zurueck(fehler=f"Zeile {zeile}: Es fehlt noch "
                                  f"{' und '.join(fehlt)}.")

        tag = parse_datum(d)
        if tag is None:
            return zurueck(fehler=f"Zeile {zeile}: Das Datum passt nicht. "
                                  "Schreib es als TT.MM.JJJJ.")
        beginn, schluss = zeit_locker(a), zeit_locker(e)
        if beginn is None or schluss is None:
            return zurueck(fehler=f"Zeile {zeile}: Start- und Endzeit "
                                  "brauchen die Form HH:MM.")
        minuten = dauer_aus_spanne(beginn, schluss)
        if not minuten:
            return zurueck(fehler=f"Zeile {zeile}: Start und Ende sind gleich – "
                                  "da kommt keine Dauer heraus.")
        if minuten > 12 * 60:
            return zurueck(fehler=f"Zeile {zeile}: Das wären {hhmm(minuten)} "
                                  "am Stück. Bitte Start und Ende prüfen.")

        # Vordefinierte Leistung und freier Text ergaenzen einander: ist
        # beides ausgefuellt, steht die einheitliche Bezeichnung vorn und
        # der eigene Zusatz dahinter. So geht nie eine Eingabe verloren.
        text = ": ".join(teil for teil in (l, b) if teil)
        satz = {"mitarbeiter": mitarbeiter, "datum": tag.isoformat(),
                "start": beginn, "ende": schluss,
                "klient": k or "Ohne Zuordnung",
                "beschreibung": text, "dauer_min": int(minuten)}
        satz["fingerprint"] = fingerprint(satz)
        satz["zeile"] = zeile
        satz["tag"] = tag
        saetze.append(satz)

    if not saetze:
        return zurueck(fehler="Es war keine Zeile ausgefüllt.")

    # Bis 0.8.8 wurde hier gegen den Bestand auf Dubletten geprueft, mit
    # einem Haken zum Uebergehen. Der ist auf Timos Wunsch entfallen: wer
    # von Hand erfasst, weiss, was er tut, und derselbe Besuch am selben Tag
    # zur selben Uhrzeit kommt in der Praxis auch echt vor. Der Fingerprint
    # wird weiter mitgeschrieben - der Listenimport braucht ihn fuer seine
    # eigene Dublettenerkennung in der Vorschau.
    with db.db() as con:
        for satz in saetze:
            con.execute(
                "INSERT INTO eintrag (import_id, mitarbeiter, datum, monat, start, "
                "ende, klient, beschreibung, dauer_min, abrechenbar, fingerprint, "
                "angelegt_am) VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?)",
                (mitarbeiter, satz["datum"], satz["tag"].strftime("%Y-%m"),
                 satz["start"], satz["ende"], satz["klient"], satz["beschreibung"],
                 satz["dauer_min"],
                 0 if norm(satz["klient"]) in NICHT_ABRECHENBAR else 1,
                 satz["fingerprint"], jetzt()))

    gesamt = sum(satz["dauer_min"] for satz in saetze)
    if len(saetze) == 1:
        return zurueck(hinweis=f"{hhmm(gesamt)} für {saetze[0]['klient']} "
                               "gespeichert.")
    return zurueck(hinweis=f"{len(saetze)} Einträge gespeichert · "
                           f"{hhmm(gesamt)} gesamt.")


# --- Auswertung -------------------------------------------------------------

@app.get("/auswertung", response_class=HTMLResponse)
def auswertung(request: Request, von_jahr: str = "", von_monat: str = "",
               bis_jahr: str = "", bis_monat: str = "",
               mitarbeiter: list[str] = Query([]),
               klient: list[str] = Query([]), q: str = "",
               nur_abrechenbar: str = ""):
    filter_ = bereichsfilter(von_jahr, von_monat, bis_jahr, bis_monat,
                             mitarbeiter, klient, q, nur_abrechenbar=nur_abrechenbar)

    with db.db() as con:
        roh = con.execute(
            f"SELECT klient, COUNT(*) n, SUM(dauer_min) m, "
            f"COUNT(DISTINCT mitarbeiter) anzahl_leute, "
            f"GROUP_CONCAT(DISTINCT mitarbeiter) leute "
            f"FROM eintrag WHERE {filter_['wo']} GROUP BY klient ORDER BY klient",
            filter_["werte"]).fetchall()
        # Zusaetzlich monatsweise: Wochenstunden und Stundensatz koennen sich
        # innerhalb des Auswertungszeitraums geaendert haben, der Verdienst
        # muss deshalb Monat fuer Monat gerechnet werden. Dieselben Zahlen
        # tragen weiter unten die Monatsbloecke.
        je_monat: dict[str, dict[str, dict]] = {}
        for r in con.execute(
                f"SELECT klient, monat, COUNT(*) n, SUM(dauer_min) m, "
                f"GROUP_CONCAT(DISTINCT mitarbeiter) leute FROM eintrag "
                f"WHERE {filter_['wo']} GROUP BY klient, monat",
                filter_["werte"]):
            je_monat.setdefault(r["klient"], {})[r["monat"]] = {
                "n": r["n"], "m": r["m"] or 0, "leute": r["leute"] or ""}
        stamm = {r["name"]: r for r in con.execute(
            "SELECT name, wochenstunden, stundensatz FROM person WHERE aktiv=1")}
        zeitraeume = zeitraeume_lesen(con)
        # Welche Monate deckt die Auswahl tatsächlich ab? Grundlage für das Soll.
        vorhandene = [r["monat"] for r in con.execute(
            f"SELECT DISTINCT monat FROM eintrag WHERE {filter_['wo']} ORDER BY monat",
            filter_["werte"])]

    if filter_["von"] and filter_["bis"]:
        # Fest umrissener Zeitraum: auch Monate ohne Daten zählen zum Soll
        monate = monatsliste(filter_["von"], filter_["bis"])
    else:
        # Offener oder monatsweiser Filter: nur die Monate, in denen etwas steht
        monate = vorhandene or [dt.date.today().strftime("%Y-%m")]

    # Bis 1.4 standen dieselben Personen in drei Boxen nebeneinander:
    # Stundentabelle, Kontingentbalken, Verdienstliste. Das erzeugte drei
    # verschieden hohe Kaesten und damit die Luecken im Raster - und man
    # musste dreimal denselben Namen suchen. Jetzt traegt eine Zeile
    # alles, was zu einer Person zu sagen ist.
    je_klient, verdienst_gesamt = [], 0.0
    gestaffelt = False
    for r in roh:
        namen = sorted({t.strip() for t in (r["leute"] or "").split(",") if t.strip()})
        p = stamm.get(r["klient"])
        zr = zeitraeume.get(r["klient"], [])
        grund_std = p["wochenstunden"] if p else 0
        grund_satz = p["stundensatz"] if p else 0

        # Soll: Monat fuer Monat mit den Werten, die in diesem Monat galten.
        # Fuer die Anzeige wird zusaetzlich festgehalten, welche
        # Wochenstunden und Saetze dabei ueberhaupt vorkamen - stehen dort
        # zwei verschiedene, waere eine einzelne Zahl in der Spalte
        # irrefuehrend.
        soll, std_stufen = 0, set()
        for monat in monate:
            std, _satz, _ = kontingent_im_monat(monat, zr, grund_std, grund_satz)
            if std:
                std_stufen.add(std)
                soll += soll_minuten(std, monat) or 0
        soll = soll or None

        # Verdienst: die Minuten JEDES Monats mit dem Satz dieses Monats.
        betrag, satz_stufen = 0.0, set()
        for monat, daten in (je_monat.get(r["klient"]) or {}).items():
            _std, satz, _ = kontingent_im_monat(monat, zr, grund_std, grund_satz)
            if satz and daten["m"]:
                satz_stufen.add(satz)
                betrag += daten["m"] / 60 * satz
        if len(std_stufen) > 1 or len(satz_stufen) > 1:
            gestaffelt = True

        # Bewusst nicht an das Soll geknuepft: mit gestaffelten Zeitraeumen
        # kann ein Monat einen Stundensatz tragen, ohne dass fuer denselben
        # Zeitraum Wochenstunden hinterlegt sind. Vorher fiel dieser
        # Verdienst stillschweigend unter den Tisch.
        verdienst_gesamt += betrag

        je_klient.append({
            "klient": r["klient"], "n": r["n"], "m": r["m"],
            "leute": namen, "anzahl_leute": r["anzahl_leute"],
            "soll": soll,
            "abweichung": (r["m"] - soll) if soll else None,
            "prozent": round(r["m"] / soll * 100) if soll else None,
            "stufen": len(std_stufen),
            "betrag": betrag,
            "satz": min(satz_stufen) if len(satz_stufen) == 1 else None,
            "saetze": sorted(satz_stufen),
        })

    # --- Monat für Monat ----------------------------------------------------
    #
    # Die Boxen oben fassen den ganzen Zeitraum zusammen. Fuer einen
    # Nachweis gegenueber dem Kostentraeger braucht es aber den einzelnen
    # Monat: was wurde geleistet, was ist daraus verdient, und mit welchem
    # Satz - der kann sich mitten im Zeitraum geaendert haben.
    #
    # Bewusst chronologisch aufsteigend: so liest sich der Block wie ein
    # Nachweis und nicht wie ein Postfach.
    #
    # Monate ohne erfasste Zeiten bleiben stehen, solange fuer sie ein Soll
    # gilt. Genau die will man sehen - eine Luecke faellt sonst nicht auf.
    monatsbloecke = []
    grundwert_monate: dict[str, int] = {}
    for monat in monate:
        zeilen, m_ist, m_soll, m_betrag, m_n = [], 0, 0, 0.0, 0
        for r in roh:
            klient = r["klient"]
            p = stamm.get(klient)
            zr = zeitraeume.get(klient, [])
            std, satz, aus_zeitraum = kontingent_im_monat(
                monat, zr, p["wochenstunden"] if p else 0,
                p["stundensatz"] if p else 0)
            daten = (je_monat.get(klient) or {}).get(monat)
            ist = daten["m"] if daten else 0
            anzahl = daten["n"] if daten else 0
            soll = soll_minuten(std, monat) or 0
            if not ist and not soll:
                # Weder gearbeitet noch beauftragt - diese Zeile traegt nichts.
                continue
            zeilenbetrag = ist / 60 * satz if (satz and ist) else 0.0
            leute = sorted({t.strip() for t in (daten["leute"] if daten else "").split(",")
                            if t.strip()})
            zeilen.append({
                "klient": klient, "n": anzahl, "m": ist, "soll": soll or None,
                "abweichung": (ist - soll) if soll else None,
                "satz": satz, "wochenstunden": std,
                "aus_zeitraum": aus_zeitraum,
                "betrag": zeilenbetrag, "leute": leute,
            })
            if not aus_zeitraum and (soll or satz):
                # Fuer die Seitenspalte: in wievielen Monaten hat der
                # Grundwert der Person gegriffen, weil kein Bescheid
                # hinterlegt ist?
                grundwert_monate[klient] = grundwert_monate.get(klient, 0) + 1
            m_ist += ist
            m_soll += soll
            m_betrag += zeilenbetrag
            m_n += anzahl
        if not zeilen:
            continue
        monatsbloecke.append({
            "monat": monat, "wort": monat_wort(monat), "zeilen": zeilen,
            "ist": m_ist, "soll": m_soll or None, "betrag": m_betrag, "n": m_n,
            "abweichung": (m_ist - m_soll) if m_soll else None,
            "prozent": round(m_ist / m_soll * 100) if m_soll else None,
            "leer": m_ist == 0,
        })

    # --- Welche Bescheide liegen dem Ganzen zugrunde? -----------------------
    # Steht in der Seitenspalte und beantwortet die Frage, die beim Lesen
    # der Zahlen als naechstes kommt: woher kommt dieser Stundensatz?
    # Nur die Zeitraeume, die den gefilterten Bereich ueberhaupt beruehren.
    filterbeginn = monatsgrenzen(monate[0])[0] if monate else ""
    filterende = monatsgrenzen(monate[-1])[1] if monate else ""
    zeitraum_liste = []
    for r in je_klient:
        treffer = [z for z in zeitraeume.get(r["klient"], [])
                   if z["von"] <= filterende
                   and (not z["bis"] or z["bis"] >= filterbeginn)]
        grund = grundwert_monate.get(r["klient"], 0)
        if treffer or grund:
            zeitraum_liste.append({
                "klient": r["klient"],
                # aufsteigend lesen, so wie die Bescheide aufeinander folgen
                "zeitraeume": list(reversed(treffer)),
                "grundwert": grund,
            })

    gesamt_ist = sum(r["m"] for r in je_klient)
    gesamt_soll = sum(b["soll"] or 0 for b in monatsbloecke)
    zusammenfassung = {
        "ist": gesamt_ist,
        "soll": gesamt_soll or None,
        "abweichung": (gesamt_ist - gesamt_soll) if gesamt_soll else None,
        "prozent": round(gesamt_ist / gesamt_soll * 100) if gesamt_soll else None,
        "betrag": verdienst_gesamt,
        "n": sum(r["n"] for r in je_klient),
        "monate": len(monate),
        "monate_mit": sum(1 for b in monatsbloecke if not b["leer"]),
        "personen": len(je_klient),
    }

    zusatz = auswahllisten()
    return templates.TemplateResponse(request=request, name="auswertung.html", context={
        "je_klient": je_klient, "verdienst_gesamt": verdienst_gesamt,
        "monate_anzahl": len(monate),
        "gesamt": sum(r["m"] for r in je_klient),
        "soll_aktiv": any(r["soll"] for r in je_klient),
        "gestaffelt": gestaffelt,
        "monatsbloecke": monatsbloecke, "zusammenfassung": zusammenfassung,
        "zeitraum_liste": zeitraum_liste,
        "zeitraum_wort": filter_["wort"], "aktive_filter": filter_["aktive"],
        "f": filter_["f"], "seite": "auswertung", **zusatz})


# --- Persönlicher Bereich -----------------------------------------------------
#
# Jede angemeldete Person sieht hier ausschliesslich die eigenen Zahlen.
# Bewusst nicht an eine Bereichsberechtigung geknuepft: es sind die eigenen
# Daten, die darf jeder sehen. Wer keinem Mitarbeiter zugeordnet ist, bekommt
# stattdessen eine Erklaerung.

def mitarbeiter_zu_benutzer(con, benutzer):
    """Welcher Mitarbeiter gehoert zum angemeldeten Konto?

    Gleiche Regel wie beim E-Mail-Versand (siehe mail.adresse_fuer):
    ausdrueckliche Zuordnung hat Vorrang, sonst Namensgleichheit.
    """
    zuordnung = ""
    try:
        zuordnung = (benutzer["mitarbeiter"] or "").strip()
    except (IndexError, KeyError, TypeError):
        # Aeltere Sitzung, die die Spalte noch nicht mitgelesen hat -
        # dann greift unten der Weg ueber die Namensgleichheit.
        zuordnung = ""
    if zuordnung:
        r = con.execute("SELECT * FROM mitarbeiter WHERE LOWER(TRIM(name))=LOWER(?)",
                        (zuordnung,)).fetchone()
        if r:
            return r
        # Zuordnung zeigt auf einen Namen, den es im Team nicht mehr gibt -
        # Zeiten koennen trotzdem noch vorhanden sein.
        return {"name": zuordnung, "monatsstunden": 0, "aktiv": 0,
                "abgabepflicht": 0, "verwaist": True}
    return con.execute(
        "SELECT * FROM mitarbeiter WHERE LOWER(name)=LOWER(?)",
        (benutzer["benutzername"],)).fetchone()


# Wieviele eigene Zeiten "Mein Bereich" hoechstens auf einmal zeigt, wenn
# kein Monat gewaehlt ist. Die Liste ist zum Nachbessern gedacht, nicht als
# zweite Uebersicht - wer weiter zurueck will, waehlt den Monat.
MEINE_ZEITEN_MAX = 300


@app.get("/meinbereich", response_class=HTMLResponse)
def meinbereich(request: Request, alle: str = "", hinweis: str = "",
                fehler: str = "", zeiten: str = ""):
    benutzer = request.state.benutzer
    with db.db() as con:
        person = mitarbeiter_zu_benutzer(con, benutzer)
        if not person:
            # Auch ohne Mitarbeiterzuordnung: der Hinweis auf fehlende
            # Bewilligungen gilt dem Team, nicht der einzelnen Person.
            return templates.TemplateResponse(
                request=request, name="meinbereich.html",
                context={"seite": "meinbereich", "person": None,
                         "monate": [], "benutzer": benutzer,
                         "spruch": spruch(), "eigene_aufgaben": [],
                         "bewilligungen": [
                             b for b in (bewilligungen_pruefen(con)
                                         if auth.darf_bewilligungen_sehen(benutzer)
                                         else [])
                             if b["art"] != "grundwert"],
                         "bewilligungen_grundwert": [
                             b for b in (bewilligungen_pruefen(con)
                                         if auth.darf_bewilligungen_sehen(benutzer)
                                         else [])
                             if b["art"] == "grundwert"],
                         "hinweis": hinweis, "fehler": fehler})

        name = person["name"]
        soll_std = float(person["monatsstunden"] or 0)
        zeilen = con.execute(
            "SELECT monat, COUNT(*) n, COALESCE(SUM(dauer_min),0) m "
            "FROM eintrag WHERE mitarbeiter=? GROUP BY monat ORDER BY monat",
            (name,)).fetchall()
        offene_vorgaenge = con.execute(
            "SELECT COUNT(*) c FROM vorgang WHERE LOWER(TRIM(zustaendig))=LOWER(?) "
            "AND status NOT IN ('Erledigt','Abgebrochen')", (name,)).fetchone()["c"]
        ueberfaellig = con.execute(
            "SELECT COUNT(*) c FROM vorgang WHERE LOWER(TRIM(zustaendig))=LOWER(?) "
            "AND status NOT IN ('Erledigt','Abgebrochen') AND frist <> '' "
            "AND frist < ?", (name, dt.date.today().isoformat())).fetchone()["c"]
        # Nicht nur zaehlen, sondern zeigen: die naechsten eigenen
        # Aufgaben stehen mit Titel, betreuter Person und Frist da.
        # Fristlose ganz nach hinten - sonst stuenden sie vor allem, was
        # wirklich draengt.
        eigene_aufgaben = con.execute(
            "SELECT id, titel, klient, art, status, prioritaet, frist "
            "FROM vorgang WHERE LOWER(TRIM(zustaendig))=LOWER(?) "
            "AND status NOT IN ('Erledigt','Abgebrochen') "
            "ORDER BY CASE WHEN frist IS NULL OR frist='' THEN 1 ELSE 0 END, "
            "frist, id LIMIT 6", (name,)).fetchall()

        # Urlaub: gezaehlt werden Kalendertage, an denen ein Eintrag steht,
        # dessen Beschreibung mit "Urlaub" beginnt. Bewusst DISTINCT nach
        # Datum, damit mehrere Zeilen am selben Tag nur einen Tag ergeben.
        # "beginnt mit" statt "enthaelt", damit z.B. eine Notiz wie
        # "Strukturplan Urlaub" nicht faelschlich als Urlaubstag zaehlt.
        urlaubsjahre = {r["jahr"]: r["tage"] for r in con.execute(
            "SELECT substr(datum,1,4) jahr, COUNT(DISTINCT datum) tage "
            "FROM eintrag WHERE mitarbeiter=? AND beschreibung LIKE 'Urlaub%' "
            "GROUP BY jahr", (name,))}

        # Die eigenen Zeiten. Bewusst ohne jede Bereichspruefung: sie
        # gehoeren dem angemeldeten Konto, und "Mein Bereich" ist die eine
        # Seite, die jeder sehen darf. Der Monatsfilter ist nur Bequem-
        # lichkeit - "alle" zeigt alles, gedeckelt auf MEINE_ZEITEN_MAX.
        # Fehlende und auslaufende Bewilligungen - dieselbe Rechnung wie
        # in den Einstellungen (main.bewilligungslage). Steht hier, weil
        # "Mein Bereich" die Seite ist, die jeder taeglich sieht: ein
        # Folgeantrag faellt sonst erst auf, wenn der Bescheid weg ist.
        alle_lagen = (bewilligungen_pruefen(con)
                      if auth.darf_bewilligungen_sehen(benutzer) else [])

        zeitmonate = [r["monat"] for r in con.execute(
            "SELECT DISTINCT monat FROM eintrag WHERE mitarbeiter=? "
            "ORDER BY monat DESC", (name,))]
        if zeiten == "alle":
            gewaehlter_monat = "alle"
        elif zeiten in zeitmonate:
            gewaehlter_monat = zeiten
        else:
            gewaehlter_monat = zeitmonate[0] if zeitmonate else "alle"
        if gewaehlter_monat == "alle":
            eigene_zeiten = con.execute(
                "SELECT * FROM eintrag WHERE mitarbeiter=? "
                "ORDER BY datum DESC, start DESC, id DESC LIMIT ?",
                (name, MEINE_ZEITEN_MAX + 1)).fetchall()
        else:
            eigene_zeiten = con.execute(
                "SELECT * FROM eintrag WHERE mitarbeiter=? AND monat=? "
                "ORDER BY datum DESC, start DESC, id DESC",
                (name, gewaehlter_monat)).fetchall()
        zeiten_gekappt = len(eigene_zeiten) > MEINE_ZEITEN_MAX
        eigene_zeiten = list(eigene_zeiten[:MEINE_ZEITEN_MAX])
        zeiten_summe = sum(z["dauer_min"] or 0 for z in eigene_zeiten)

    ist_je_monat = {r["monat"]: {"m": r["m"], "n": r["n"]} for r in zeilen}
    dieser_monat = dt.date.today().strftime("%Y-%m")

    # Von der ersten erfassten Zeit bis heute jeden Monat auffuellen, damit
    # ein Monat ohne Eintraege sichtbar als Luecke erscheint statt zu fehlen.
    monate = []
    if ist_je_monat:
        start = min(ist_je_monat)
        lauf = dt.date(int(start[:4]), int(start[5:7]), 1)
        ende = dt.date.today().replace(day=1)
        while lauf <= ende:
            schluessel = lauf.strftime("%Y-%m")
            daten = ist_je_monat.get(schluessel, {"m": 0, "n": 0})
            soll_min = int(round(soll_std * 60))
            monate.append({
                "monat": schluessel,
                "wort": monat_wort(schluessel),
                "ist": daten["m"],
                "n": daten["n"],
                "soll": soll_min,
                "saldo": daten["m"] - soll_min if soll_min else 0,
                "laufend": schluessel == dieser_monat,
            })
            lauf = (lauf + dt.timedelta(days=32)).replace(day=1)
        monate.reverse()

    # Der laufende Monat ist noch nicht vorbei - er wuerde den Saldo
    # kuenstlich ins Minus ziehen und wird deshalb getrennt ausgewiesen.
    abgeschlossen = [m for m in monate if not m["laufend"]]
    gesamt = {
        "ist": sum(m["ist"] for m in abgeschlossen),
        "soll": sum(m["soll"] for m in abgeschlossen),
        "monate": len(abgeschlossen),
    }
    gesamt["saldo"] = gesamt["ist"] - gesamt["soll"]
    laufend = next((m for m in monate if m["laufend"]), None)

    # Der zuletzt abgeschlossene Monat und der Schnitt der letzten drei -
    # daraus laesst sich ablesen, ob man gerade regelmaessig ueber oder
    # unter dem Soll liegt. Ein ueber Jahre aufsummierter Gesamtsaldo waere
    # dagegen kaum interpretierbar.
    letzter = abgeschlossen[0] if abgeschlossen else None
    dreimonate = abgeschlossen[:3]
    trend = None
    if dreimonate and soll_std:
        schnitt_ist = sum(m["ist"] for m in dreimonate) / len(dreimonate)
        schnitt_soll = sum(m["soll"] for m in dreimonate) / len(dreimonate)
        trend = {
            "monate": len(dreimonate),
            "schnitt": int(round(schnitt_ist)),
            "saldo": int(round(schnitt_ist - schnitt_soll)),
            "reihe": list(reversed(dreimonate)),
        }

    if not alle:
        monate = monate[:13]

    # --- Diagrammdaten ------------------------------------------------------
    # Bewusst als fertig gerechnetes SVG statt einer Diagramm-Bibliothek:
    # das Projekt laedt keine externen Skripte, und ein Balkendiagramm
    # braucht keine.
    letzte = list(reversed(monate[:12]))  # chronologisch, aelteste links
    diagramm = None
    if letzte:
        breite, hoehe = 420, 190
        oben, unten, links, rechts = 14, 32, 34, 8
        flaeche = hoehe - oben - unten
        spalte = (breite - links - rechts) / max(len(letzte), 1)
        soll_min = int(round(soll_std * 60))
        spitze = max([m["ist"] for m in letzte] + [soll_min, 60])
        spitze = spitze * 1.18

        balken, saldopunkte, summe = [], [], 0
        for i, m in enumerate(letzte):
            h = flaeche * (m["ist"] / spitze)
            x = links + i * spalte + spalte * 0.22
            balkenbreite = spalte * 0.56
            erreicht = soll_min and m["ist"] >= soll_min
            balken.append({
                "x": round(x, 1), "y": round(oben + flaeche - h, 1),
                "b": round(balkenbreite, 1), "h": round(max(h, 1), 1),
                "mitte": round(x + balkenbreite / 2, 1),
                "kurz": MONATSNAMEN.get(m["monat"][5:7], "")[:3],
                "jahr": m["monat"][:4],
                "wert": hhmm(m["ist"]),
                "wort": m["wort"],
                "klasse": ("balken-laufend" if m["laufend"] else
                           "balken-gut" if erreicht else
                           "balken-unter" if soll_min else "balken-neutral"),
            })
            if soll_min and not m["laufend"]:
                summe += m["saldo"]
                saldopunkte.append({"x": round(x + balkenbreite / 2, 1),
                                    "saldo": summe})

        # Der Saldoverlauf bekommt eine eigene Skala, sonst waere er neben
        # den Monatsbalken kaum zu erkennen.
        linie = ""
        if len(saldopunkte) > 1:
            grenze = max(abs(p["saldo"]) for p in saldopunkte) or 1
            mitte = oben + flaeche / 2
            teile = []
            for p in saldopunkte:
                y = mitte - (p["saldo"] / grenze) * (flaeche / 2 * 0.82)
                teile.append(f"{p['x']},{round(y, 1)}")
            linie = " ".join(teile)

        diagramm = {
            "breite": breite, "hoehe": hoehe, "oben": oben, "links": links,
            "flaeche": flaeche, "grundlinie": oben + flaeche,
            "balken": balken, "linie": linie,
            "soll_y": round(oben + flaeche - flaeche * (soll_min / spitze), 1)
                      if soll_min else None,
            "soll_wert": hhmm(soll_min) if soll_min else None,
            "rechts_x": breite - rechts,
            "mittellinie": round(oben + flaeche / 2, 1),
        }

    # --- Urlaub -------------------------------------------------------------
    jahr = dt.date.today().strftime("%Y")
    anspruch = float(person["urlaubstage"] or 0)
    genommen = urlaubsjahre.get(jahr, 0)
    urlaub = {
        "jahr": jahr,
        "anspruch": anspruch,
        "genommen": genommen,
        "rest": anspruch - genommen,
        "anteil": min(100, round(genommen / anspruch * 100)) if anspruch else 0,
        "ueberzogen": bool(anspruch and genommen > anspruch),
        "vorjahre": sorted(
            ((j, t) for j, t in urlaubsjahre.items() if j != jahr),
            reverse=True)[:3],
    }

    return templates.TemplateResponse(
        request=request, name="meinbereich.html", context={
            "seite": "meinbereich", "person": person, "name": person["name"],
            "soll_std": soll_std, "monate": monate, "gesamt": gesamt,
            "laufend": laufend, "alle": bool(alle), "benutzer": benutzer,
            "hinweis": hinweis, "fehler": fehler,
            "diagramm": diagramm, "urlaub": urlaub,
            "letzter": letzter, "trend": trend,
            "offene_vorgaenge": offene_vorgaenge, "ueberfaellig": ueberfaellig,
            "eigene_aufgaben": eigene_aufgaben, "spruch": spruch(),
            "heute": dt.date.today().isoformat(),
            "bewilligungen": [b for b in alle_lagen if b["art"] != "grundwert"],
            "bewilligungen_grundwert": [b for b in alle_lagen
                                        if b["art"] == "grundwert"],
            "eigene_zeiten": eigene_zeiten, "zeitmonate": zeitmonate,
            "zeiten_monat": gewaehlter_monat, "zeiten_gekappt": zeiten_gekappt,
            "zeiten_summe": zeiten_summe, "zeiten_max": MEINE_ZEITEN_MAX,
            "verwaist": isinstance(person, dict) and person.get("verwaist")})


# --- Ideen und Changelog ----------------------------------------------------

IDEEN_ARTEN = ["Idee", "Kritik", "Fehler", "Frage"]


def bloecke_lesen(pfad: str) -> list[dict]:
    """Liest eine Datei im Format '## Kopfzeile' plus folgenden Textzeilen.

    Gibt die Blöcke in Dateireihenfolge zurück, jeweils mit Kopfzeile und
    den restlichen Zeilen als Text.
    """
    try:
        with open(pfad, encoding="utf-8") as f:
            roh = f.read()
    except OSError:
        return []

    # Kommentarzeilen ganz am Zeilenanfang mit einzelnem # entfernen,
    # damit der Dateikopf nicht als Block auftaucht. ## bleibt Trenner.
    roh = "\n".join(z for z in roh.splitlines()
                    if not (z.startswith("#") and not z.startswith("##")))

    ergebnis = []
    for teil in roh.split("\n## "):
        teil = teil.strip()
        if not teil:
            continue
        if teil.startswith("## "):
            teil = teil[3:]
        zeilen = teil.splitlines()
        ergebnis.append({"kopf": zeilen[0].strip(),
                         "text": "\n".join(zeilen[1:]).strip()})
    return ergebnis


def idee_saeubern(art: str, wer: str, text: str):
    """Prüft und entschärft eine Eingabe. Gibt (art, wer, text, fehler) zurück."""
    text = text.strip()
    if not text:
        return None, None, None, "Da war noch nichts geschrieben."
    if len(text) > 4000:
        return None, None, None, "Bitte auf 4000 Zeichen kürzen."

    art = art if art in IDEEN_ARTEN else "Idee"
    wer = re.sub(r"[|\n\r]", " ", wer).strip()[:60] or "anonym"
    # Zeilen, die wie eine Blocktrennung aussehen, entschärfen
    text = "\n".join(
        ("#\u200b# " + z.lstrip("# ").rstrip()) if z.lstrip().startswith("##") else z
        for z in text.splitlines())
    return art, wer, text, None


def ideen_kopfzeilen() -> list[str]:
    """Die Kommentarzeilen am Dateianfang, damit sie beim Umschreiben bleiben."""
    try:
        with open(IDEEN_DATEI, encoding="utf-8") as f:
            zeilen = f.read().splitlines()
    except OSError:
        return []
    kopf = []
    for z in zeilen:
        if z.lstrip().startswith("##"):
            break
        kopf.append(z)
    return kopf


def ideen_laden() -> list[dict]:
    """Alle Einträge in Dateireihenfolge, mit Nummer als Schlüssel."""
    gesammelt = []
    for nr, b in enumerate(bloecke_lesen(IDEEN_DATEI)):
        # Kopfzeile: "17.08.2026 09:45 | Idee | Name"
        teile = [t.strip() for t in b["kopf"].split("|")]
        gesammelt.append({
            "nr": nr,
            "zeit": teile[0] if teile else "",
            "art": teile[1] if len(teile) > 1 else "Idee",
            "wer": teile[2] if len(teile) > 2 else "",
            "text": b["text"],
        })
    return gesammelt


def ideen_schreiben(eintraege: list[dict]) -> str | None:
    """Schreibt die ganze Datei neu. Gibt eine Fehlermeldung zurück oder None."""
    text = "\n".join(ideen_kopfzeilen()).rstrip()
    for e in eintraege:
        text += (f"\n\n## {e['zeit']} | {e['art']} | {e['wer']}\n{e['text']}")
    try:
        os.makedirs(os.path.dirname(IDEEN_DATEI) or ".", exist_ok=True)
        with open(IDEEN_DATEI, "w", encoding="utf-8") as f:
            f.write(text.lstrip("\n") + "\n")
    except OSError as e:
        return f"Konnte nicht gespeichert werden: {e}"
    return None


def ideen_zurueck(**werte):
    return RedirectResponse("/ideen?" + urlencode(werte) if werte else "/ideen",
                            status_code=303)


@app.get("/ideen", response_class=HTMLResponse)
def ideen(request: Request, hinweis: str = "", fehler: str = "",
          bearbeiten: int = -1):
    eintraege = ideen_laden()
    eintraege.reverse()  # neueste zuerst, Nummer bleibt die aus der Datei
    return templates.TemplateResponse(request=request, name="ideen.html", context={
        "eintraege": eintraege, "arten": IDEEN_ARTEN, "hinweis": hinweis,
        "fehler": fehler, "datei": IDEEN_DATEI, "bearbeiten": bearbeiten,
        "seite": "ideen"})


@app.post("/ideen")
def idee_anlegen(art: str = Form("Idee"), wer: str = Form(""),
                 text: str = Form("")):
    art, wer, text, fehler = idee_saeubern(art, wer, text)
    if fehler:
        return ideen_zurueck(fehler=fehler)

    satz = (f"\n## {dt.datetime.now().strftime('%d.%m.%Y %H:%M')} | {art} | {wer}\n"
            f"{text}\n")
    try:
        os.makedirs(os.path.dirname(IDEEN_DATEI) or ".", exist_ok=True)
        with open(IDEEN_DATEI, "a", encoding="utf-8") as f:
            f.write(satz)
    except OSError as e:
        return ideen_zurueck(fehler=f"Konnte nicht gespeichert werden: {e}")

    return ideen_zurueck(hinweis="Danke, ist notiert.")


@app.post("/ideen/{nr}/bearbeiten")
def idee_bearbeiten(nr: int, art: str = Form("Idee"), wer: str = Form(""),
                    text: str = Form("")):
    eintraege = ideen_laden()
    if not 0 <= nr < len(eintraege):
        return ideen_zurueck(fehler="Der Eintrag existiert nicht mehr.")

    art, wer, text, fehler = idee_saeubern(art, wer, text)
    if fehler:
        return ideen_zurueck(fehler=fehler, bearbeiten=nr)

    eintraege[nr].update({"art": art, "wer": wer, "text": text})
    problem = ideen_schreiben(eintraege)
    if problem:
        return ideen_zurueck(fehler=problem)
    return ideen_zurueck(hinweis="Eintrag geändert.")


@app.post("/ideen/{nr}/loeschen")
def idee_loeschen(nr: int):
    eintraege = ideen_laden()
    if not 0 <= nr < len(eintraege):
        return ideen_zurueck(fehler="Der Eintrag existiert nicht mehr.")
    weg = eintraege.pop(nr)
    problem = ideen_schreiben(eintraege)
    if problem:
        return ideen_zurueck(fehler=problem)
    return ideen_zurueck(hinweis=f"Eintrag von {weg['wer']} entfernt.")


# --- Das eigene Konto ---------------------------------------------------------
#
# Die Benutzerverwaltung unter Einstellungen ist Administratoren vorbehalten
# (auth.ADMIN_NUR_PFADE). Damit ein normales Konto trotzdem sein Passwort
# wechseln kann, sitzt hier die kleine Selbstbedienung: eigene E-Mail-Adresse
# und eigenes Passwort, sonst nichts. Rolle, Bereiche und die Zuordnung zu
# einem Mitarbeiter bleiben ausdruecklich Sache der Administration - sonst
# koennte sich jeder selbst hochstufen.
#
# Der Pfad liegt unter /meinbereich und damit ausserhalb von
# ADMIN_NUR_PFADE und ausserhalb jedes Bereichs: es sind die eigenen Daten.

def _konto_zurueck(**werte):
    return RedirectResponse("/meinbereich?" + urlencode(werte), status_code=303)


@app.post("/meinbereich/konto")
def konto_speichern(request: Request, email: str = Form(""),
                    passwort_alt: str = Form(""), passwort_neu: str = Form(""),
                    passwort_neu2: str = Form("")):
    benutzer = request.state.benutzer
    email = email.strip()
    meldungen = []

    with db.db() as con:
        satz = con.execute("SELECT * FROM benutzer WHERE id=?",
                           (benutzer["id"],)).fetchone()
        if satz is None:
            return _konto_zurueck(fehler="Dieses Konto gibt es nicht mehr.")

        if passwort_neu or passwort_neu2 or passwort_alt:
            # Das aktuelle Passwort ist Pflicht. Sonst koennte jemand an
            # einem unbeaufsichtigt offenen Bildschirm das Konto uebernehmen
            # und die eigentliche Inhaberin aussperren.
            if not db.passwort_pruefen(passwort_alt, satz["passwort_hash"]):
                return _konto_zurueck(fehler="Das aktuelle Passwort stimmt nicht.")
            if len(passwort_neu) < 8:
                return _konto_zurueck(
                    fehler="Das neue Passwort braucht mindestens acht Zeichen.")
            if passwort_neu != passwort_neu2:
                return _konto_zurueck(
                    fehler="Die beiden neuen Passwörter sind nicht gleich.")
            con.execute("UPDATE benutzer SET passwort_hash=? WHERE id=?",
                        (db.passwort_hashen(passwort_neu), benutzer["id"]))
            # Alle anderen Sitzungen beenden. Wer das Passwort wechselt,
            # will in aller Regel genau das - die eigene bleibt bestehen.
            token = request.cookies.get(auth.COOKIE_NAME, "")
            con.execute("DELETE FROM sitzung WHERE benutzer_id=? AND token<>?",
                        (benutzer["id"], token))
            meldungen.append("Passwort geändert")

        if email != (satz["email"] or ""):
            con.execute("UPDATE benutzer SET email=? WHERE id=?",
                        (email or None, benutzer["id"]))
            meldungen.append("E-Mail-Adresse gespeichert" if email
                             else "E-Mail-Adresse entfernt")

    if not meldungen:
        return _konto_zurueck(hinweis="Es gab nichts zu ändern.")
    return _konto_zurueck(hinweis=" · ".join(meldungen) + ".")


@app.get("/changelog", response_class=HTMLResponse)
def changelog(request: Request):
    # Neueste Version oben, die Liste selbst ist chronologisch gepflegt
    return templates.TemplateResponse(request=request, name="changelog.html", context={
        "staende": list(reversed(CHANGELOG)), "seite": "changelog"})


# --- Export -----------------------------------------------------------------

SPALTEN = ["Datum", "Betreuter", "Beginn", "Ende", "Dauer",
           "Leistung/Beschreibung", "Mitarbeiter"]


def hole_zeilen(wo, werte):
    with db.db() as con:
        return con.execute(
            f"SELECT * FROM eintrag WHERE {wo} ORDER BY datum, mitarbeiter, start",
            werte).fetchall()


def baue_xlsx(zeilen) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Nachweis"
    ws.append(SPALTEN)
    for zelle in ws[1]:
        zelle.font = Font(bold=True, color="FFFFFF")
        zelle.fill = PatternFill("solid", fgColor="272827")
        zelle.alignment = Alignment(vertical="center")
    for z in zeilen:
        ws.append([deutsch(z["datum"]), z["klient"], z["start"], z["ende"],
                   hhmm(z["dauer_min"]), z["beschreibung"], z["mitarbeiter"]])
    summe = sum(z["dauer_min"] for z in zeilen)
    ws.append([])
    ws.append(["Gesamt", "", "", "", hhmm(summe), f"{len(zeilen)} Einträge", ""])
    for zelle in ws[ws.max_row]:
        zelle.font = Font(bold=True)
    for spalte, breite in zip("ABCDEFG", [12, 24, 9, 9, 10, 54, 16]):
        ws.column_dimensions[spalte].width = breite
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(2, len(zeilen) + 1)}"

    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer


def exportname(filter_: dict, endung: str) -> str:
    """Dateiname aus dem gewählten Zeitraum und Mitarbeiter."""
    von, bis = filter_["von"], filter_["bis"]
    nur_a, nur_b = filter_.get("nur_monate", ("", ""))
    if von and bis and von == bis:
        zeit = von
    elif von and bis:
        zeit = f"{von}_bis_{bis}"
    elif von:
        zeit = f"ab_{von}"
    elif bis:
        zeit = f"bis_{bis}"
    elif nur_a or nur_b:
        # Monat ohne Jahreszahl: Monatsname statt Datumsspanne
        namen = [MONATSNAMEN.get(x, x) for x in (nur_a, nur_b) if x]
        zeit = ("-".join(dict.fromkeys(namen))) + "_alle-Jahre"
    else:
        zeit = "gesamt"
    teile = ["Zeitnachweis", zeit]
    if filter_["f"]["mitarbeiter"]:
        teile.append(sicherer_name(filter_["f"]["mitarbeiter"]).replace(" ", "-"))
    return "_".join(teile) + endung


@app.get("/export.xlsx")
def export_xlsx(von_jahr: str = "", von_monat: str = "", bis_jahr: str = "",
                bis_monat: str = "", mitarbeiter: list[str] = Query([]),
                klient: list[str] = Query([]),
                q: str = "", import_id: int = 0, nur_abrechenbar: str = ""):
    filter_ = bereichsfilter(von_jahr, von_monat, bis_jahr, bis_monat,
                             mitarbeiter, klient, q, import_id, nur_abrechenbar)
    puffer = baue_xlsx(hole_zeilen(filter_["wo"], filter_["werte"]))
    name = exportname(filter_, ".xlsx")
    return StreamingResponse(
        puffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'})


@app.get("/export.csv")
def export_csv(von_jahr: str = "", von_monat: str = "", bis_jahr: str = "",
               bis_monat: str = "", mitarbeiter: list[str] = Query([]),
               klient: list[str] = Query([]),
               q: str = "", import_id: int = 0, nur_abrechenbar: str = ""):
    import csv as csvmod
    filter_ = bereichsfilter(von_jahr, von_monat, bis_jahr, bis_monat,
                             mitarbeiter, klient, q, import_id, nur_abrechenbar)
    puffer = io.StringIO()
    schreiber = csvmod.writer(puffer, delimiter=";")
    schreiber.writerow(SPALTEN)
    for z in hole_zeilen(filter_["wo"], filter_["werte"]):
        schreiber.writerow([deutsch(z["datum"]), z["klient"], z["start"], z["ende"],
                            hhmm(z["dauer_min"]), z["beschreibung"],
                            z["mitarbeiter"]])
    daten = ("\ufeff" + puffer.getvalue()).encode("utf-8")
    name = exportname(filter_, ".csv")
    return StreamingResponse(io.BytesIO(daten), media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{name}"'})

@app.get("/gesundheit")
def gesundheit():
    with db.db() as con:
        con.execute("SELECT 1")
    return {"status": "ok", "zeit": jetzt(), "version": VERSION}

# --- Verwaltungsvorgaenge ---------------------------------------------------
#
# Das Modul steckt in vorgaenge.py. Erst hier eingebunden, damit die Templates
# und Filter oben schon fertig eingerichtet sind.

from . import vorgaenge as _vorgaenge  # noqa: E402

_vorgaenge.setup(templates)
app.include_router(_vorgaenge.router)


# --- Einstellungen ----------------------------------------------------------
#
# Ebenfalls ein eigenes Modul (einstellungen.py). Es bekommt hier die Werte
# gereicht, die es aus main.py braucht - so gibt es keinen Ringschluss beim
# Import, und main.py bleibt die einzige Stelle, an der Pfade und
# Umgebungsvariablen gelesen werden.

from . import einstellungen as _einstellungen  # noqa: E402

_einstellungen.setup(templates, {
    "jetzt": jetzt,
    "heute": lambda: dt.date.today().isoformat(),
    "bewilligungslage": bewilligungslage,
    "BEWILLIGUNG_HANDLUNG": BEWILLIGUNG_HANDLUNG,
    "monat_wort": monat_wort,
    "VERSION": VERSION,
    "MAX_UPLOAD_MB": MAX_UPLOAD_MB,
    "WECKER_INTERVALL": WECKER_INTERVALL,
    "SPRUCH_DATEI": SPRUCH_DATEI,
    "IDEEN_DATEI": IDEEN_DATEI,
    "STRINGS_DATEI": STRINGS_DATEI,
})
app.include_router(_einstellungen.router)


# --- Fuhrpark ---------------------------------------------------------------
#
# Fahrzeuge, ihre Ereignisse und die Auswertung stecken in kfz.py. Dasselbe
# Muster wie oben: eigener Router, eingebunden erst hier, wenn die Templates
# stehen. Das Modul braucht nichts aus main.py ausser den Templates.

from . import kfz as _kfz  # noqa: E402

_kfz.setup(templates)
app.include_router(_kfz.router)


# --- Wiki -------------------------------------------------------------------
#
# Die Wissensbasis liegt als Markdown-Dateien im Ordner /wiki, das Modul
# steckt in wiki.py. Wie bei einstellungen.py bekommt es den Pfad hier
# gereicht, damit main.py die einzige Stelle bleibt, an der
# Umgebungsvariablen gelesen werden.

_wiki.setup(templates, {"WIKI_PFAD": WIKI_PFAD})
app.include_router(_wiki.router)


# --- Dateien ----------------------------------------------------------------
#
# Bilder, PDFs und Office-Dateien im Ordner /files, damit sich im Wiki
# etwas verlinken laesst, ohne die Datei ueber die Freigabe dorthin legen
# zu muessen. Wie beim Wiki ist der Ordner selbst der Bestand - keine
# Tabelle, damit derselbe Ordner auch ueber die Dateifreigabe nutzbar
# bleibt.

from . import dateien as _dateien  # noqa: E402

_dateien.setup(templates, {"FILES_PFAD": FILES_PFAD,
                           "MAX_UPLOAD_MB": MAX_UPLOAD_MB})
app.include_router(_dateien.router)
